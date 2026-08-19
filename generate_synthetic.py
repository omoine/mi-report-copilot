"""Generate a month of synthetic intraday liquidity data.

Shape comes from the OCR profile (currency mix, status mix, amount magnitudes);
identifiers are invented, because the OCR could not recover a single real one.

What this fixes against the original sample, all of which were recorded as gaps
in DATA_REQUIREMENTS.md:

- a real intraday time shape, so an hourly profile has mass in each bucket
- a month of business days, so day-of-week and month-end effects exist
- realistic cardinality: thousands of account/currency pairs, not two dozen
- a display-currency amount on the transfer view, so cross-currency totals mean
  something
- a shared key across the views, so they can actually be combined
- deliberate, findable anomalies, so a report can show something worth seeing

Usage:
    python generate_synthetic.py --out data/synthetic_liquidity_month.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Reference data. Rates are invented and flat across the period on purpose:
# a moving rate would imply an FX time series the tool cannot source.
# --------------------------------------------------------------------------
FX_TO_GBP = {
    "GBP": 1.0, "EUR": 0.858, "USD": 0.784, "JPY": 0.00525, "CHF": 0.897,
    "HKD": 0.1005, "SGD": 0.583, "AUD": 0.513, "NOK": 0.0735, "SEK": 0.073,
    "ZAR": 0.043, "CNY": 0.109, "INR": 0.0094, "CAD": 0.574, "NZD": 0.472,
    "DKK": 0.115, "PLN": 0.198, "MXN": 0.0428, "TRY": 0.0231, "AED": 0.213,
    "SAR": 0.209, "THB": 0.0221, "KRW": 0.00058, "TWD": 0.0243, "BRL": 0.142,
    "CZK": 0.0339, "HUF": 0.00215, "ILS": 0.211, "RON": 0.172,
}

VENUE_STEMS = ["Alder", "Bramble", "Cedar", "Dunlin", "Elmwood", "Fernway",
               "Granite", "Harrow", "Ironbark", "Juniper", "Kestrel", "Larkspur",
               "Mallow", "Nightjar", "Oriel", "Pinehurst", "Redstart", "Sablewood",
               "Thornbury", "Umberton", "Vetchling", "Wrenfield", "Yarrow", "Zephyr"]
VENUE_KIND = ["Bank", "Trust", "Clearing", "Settlement Bank", "Commercial Bank"]
VENUE_CITY = ["London", "New York", "Frankfurt", "Zurich", "Singapore", "Tokyo",
              "Hong Kong", "Sydney", "Luxembourg", "Toronto", "Mexico City",
              "Warsaw", "Stockholm", "Oslo", "Dublin", "Paris", "Madrid", "Milan"]
ENTITY_STEMS = ["Alder", "Cedar", "Granite", "Juniper", "Kestrel", "Oriel",
                "Pinehurst", "Thornbury", "Wrenfield"]
ENTITY_FORM = ["plc", "N.A.", "S.A.", "Ltd", "AG", "Pte Ltd"]
FIRST = ["Avery", "Blake", "Casey", "Devon", "Emery", "Finley", "Greer", "Harper",
         "Jules", "Kai", "Lennox", "Marlow", "Noor", "Oakley", "Quinn", "Reese",
         "Sawyer", "Tatum", "Wren", "Ellis"]
LAST = ["Ashford", "Brookes", "Calloway", "Delaney", "Ellison", "Fairweather",
        "Gallagher", "Hollis", "Jennings", "Kirby", "Lindqvist", "Merrick",
        "Novak", "Osgood", "Prentice", "Rhodes", "Sinclair", "Thorne"]
DESKS = ["BRAA", "BRBX", "BRCC", "BRGL", "BRGX", "BRYG", "BRMD", "BRNK"]
BOOK_SUFFIX = ["", "TRSY", "CASH", "MM"]
CASHFLOW_TYPES = ["TCX", "TCC", "TCG", "TCQ", "TCP"]     # anonymised type codes
NET_TYPES = ["NETTER", "GROSS", "INTERNAL"]
TRADE_STATUSES = ["NEW", "AMEND", "CANCEL"]
ACCOUNT_PURPOSE = ["Operating", "Client Money", "Settlement", "Fees", "Liquidity",
                   "Collateral", "Nostro", "Reserve"]

# Share of the day's transfers arriving in each hour. Two peaks either side of
# lunch, a cutoff tail, and near-nothing overnight.
INTRADAY_CURVE = {
    0: .002, 1: .002, 2: .003, 3: .004, 4: .006, 5: .011, 6: .028, 7: .062,
    8: .095, 9: .118, 10: .112, 11: .088, 12: .052, 13: .058, 14: .091,
    15: .098, 16: .073, 17: .048, 18: .022, 19: .011, 20: .006, 21: .004,
    22: .003, 23: .003,
}
WEEKDAY_WEIGHT = {0: 1.06, 1: 1.10, 2: 1.02, 3: 1.00, 4: 0.92}  # Mon..Fri

# Populated during generation so the data dictionary documents the domains that
# were actually written, after anonymisation.
DOMAINS_SEEN: dict[str, list[str]] = {}


def anonymise_mix(mix: dict[str, float], category: str) -> dict[str, float]:
    """Route enum values through the golden source before they reach output.

    The OCR profile records the real domain values, and some of those are
    internal system names. Generated data is published, so it must carry the
    surrogate - otherwise the generator quietly undoes the anonymisation.
    """
    try:
        import anonymise
    except ImportError:  # pragma: no cover - generator can still run standalone
        return mix
    mapping = anonymise._load()
    out: dict[str, float] = {}
    for value, weight in mix.items():
        safe = anonymise._map_value(value, category, mapping)
        out[safe] = out.get(safe, 0.0) + weight
    anonymise._save(mapping)
    return out


def _weighted(rnd: random.Random, mix: dict[str, float]) -> str:
    keys = list(mix)
    return rnd.choices(keys, weights=[mix[k] for k in keys], k=1)[0]


def _amount_for(rnd: random.Random, digit_mix: dict[str, float]) -> float:
    """Draw a magnitude consistent with the digit-length mix seen in the OCR,
    then spread it log-uniformly inside that decade."""
    digits = int(_weighted(rnd, digit_mix)) if digit_mix else 10
    low = 10 ** max(digits - 4, 0)          # OCR digits include 3 decimal places
    high = 10 ** max(digits - 3, 1)
    return round(math.exp(rnd.uniform(math.log(low), math.log(high))), 3)


def business_days(end: dt.date, count: int) -> list[dt.date]:
    days, cursor = [], end
    while len(days) < count:
        if cursor.weekday() < 5:
            days.append(cursor)
        cursor -= dt.timedelta(days=1)
    return sorted(days)


def _timestamp(rnd: random.Random, day: dt.date) -> dt.datetime:
    hour = int(_weighted(rnd, {str(h): w for h, w in INTRADAY_CURVE.items()}))
    return dt.datetime.combine(day, dt.time(hour, rnd.randrange(60), rnd.randrange(60)))


class Universe:
    """The account population, most of which is dormant in any given month."""

    def __init__(self, rnd: random.Random, profile: dict, n_pairs: int, n_active: int):
        self.rnd = rnd
        self.currency_mix = profile.get("currency_mix") or {"GBP": 1.0}
        self.currency_mix = {c: w for c, w in self.currency_mix.items() if c in FX_TO_GBP}

        self.venues = [f"{rnd.choice(VENUE_STEMS)} {rnd.choice(VENUE_KIND)} "
                       f"{rnd.choice(VENUE_CITY)}" for _ in range(60)]
        self.venues = sorted(set(self.venues))
        self.entities = sorted({f"{rnd.choice(ENTITY_STEMS)} Group {rnd.choice(ENTITY_FORM)}"
                                for _ in range(9)})
        self.people = sorted({f"{rnd.choice(FIRST)} {rnd.choice(LAST)}" for _ in range(40)})
        self.counterparties = sorted({f"{rnd.choice(VENUE_STEMS)} {rnd.choice(VENUE_KIND)}"
                                      for _ in range(55)})

        # Account/currency pairs. Concentration is deliberate: a handful of
        # accounts carry most of the flow, which is what makes ranking and
        # concentration questions meaningful.
        self.pairs: list[dict] = []
        for i in range(n_pairs):
            ccy = _weighted(rnd, self.currency_mix)
            self.pairs.append({
                "account": f"{rnd.randrange(10**11, 10**12):012d}",
                "currency": ccy,
                "entity": rnd.choice(self.entities),
                "sort_code": f"{rnd.randrange(100000, 999999)}",
                "name": f"{rnd.choice(ACCOUNT_PURPOSE)} {ccy} {i % 90 + 1:02d}",
                "desk": rnd.choice(DESKS),
            })
        self.active = rnd.sample(self.pairs, k=min(n_active, len(self.pairs)))
        # Pareto-ish weighting over the active set.
        self.active_weights = [1 / (i + 1) ** 0.65 for i in range(len(self.active))]

    def pick_active(self) -> dict:
        return self.rnd.choices(self.active, weights=self.active_weights, k=1)[0]


def generate(profile: dict, days: int, n_pairs: int, n_active: int,
             end: dt.date, seed: int) -> dict[str, list]:
    rnd = random.Random(seed)
    uni = Universe(rnd, profile, n_pairs, n_active)
    dates = business_days(end, days)

    status_mix = profile.get("transfer_status_mix") or {"LEDGER_OR_CASHFLOW_RECEIVED": 1.0}
    strategy_mix = profile.get("sending_strategy_mix") or {"FISS": 1.0}
    message_mix = profile.get("message_status_mix") or {"FISS_ACKNOWLEDGED": 1.0}
    digit_mix = profile.get("amount_digit_mix") or {}

    # The OCR sample saw almost nothing but the happy path. Real MI needs the
    # exceptions to exist, so a floor is applied to the rare statuses.
    status_mix = dict(status_mix)
    status_mix.setdefault("PENDING_APPROVAL", 0.0)
    status_mix.setdefault("FAILED", 0.0)
    status_mix["PENDING_APPROVAL"] = max(status_mix["PENDING_APPROVAL"], 0.045)
    status_mix["FAILED"] = max(status_mix["FAILED"], 0.018)
    strategy_mix = dict(strategy_mix)
    strategy_mix.setdefault("SWIFT", 0.0)
    strategy_mix.setdefault("INTERNAL", 0.0)
    strategy_mix["SWIFT"] = max(strategy_mix["SWIFT"], 0.30)
    strategy_mix["INTERNAL"] = max(strategy_mix["INTERNAL"], 0.12)
    # Message statuses embed the messaging system's name, so derive the missing
    # ones from whichever acknowledged-status is present rather than hardcoding.
    message_mix = dict(message_mix)
    ack = next((k for k in message_mix if k.endswith("_ACKNOWLEDGED")), None)
    if ack:
        prefix = ack.rsplit("_", 1)[0]
        for suffix, floor in (("PENDING", 0.06), ("REJECTED", 0.025)):
            key = f"{prefix}_{suffix}"
            message_mix[key] = max(message_mix.get(key, 0.0), floor)

    # Anonymise the enum domains before any of this reaches the workbook.
    strategy_mix = anonymise_mix(strategy_mix, "system")
    message_mix = anonymise_mix(message_mix, "system_status")
    DOMAINS_SEEN["strategy"] = list(strategy_mix)
    DOMAINS_SEEN["message"] = list(message_mix)

    # One stressed day and one quiet day, so the month is not flat.
    spike_day = dates[len(dates) // 2] if dates else None
    quiet_day = dates[max(len(dates) - 4, 0)] if dates else None

    transfers, ledger, client = [], [], []
    ref_seq, txn_seq = 1, 1

    for day in dates:
        weekday_factor = WEEKDAY_WEIGHT.get(day.weekday(), 1.0)
        month_end = 1.35 if day.day >= 28 else 1.0
        stress = 2.1 if day == spike_day else (0.55 if day == quiet_day else 1.0)
        factor = weekday_factor * month_end * stress

        # ---------------- nostro transfers ----------------
        for _ in range(int(rnd.gauss(190, 22) * factor)):
            src, tgt = uni.pick_active(), uni.pick_active()
            ccy = src["currency"]
            amount = _amount_for(rnd, digit_mix) * (1.8 if day == spike_day else 1.0)
            created = _timestamp(rnd, day)
            status = _weighted(rnd, status_mix)
            approved_by, approved_time = None, None
            if status != "PENDING_APPROVAL":
                approved_time = created + dt.timedelta(minutes=rnd.randint(2, 95))
                approved_by = rnd.choice(uni.people)
            reference = f"NT{day:%y%m%d}{ref_seq:05d}"
            ref_seq += 1
            rate = FX_TO_GBP.get(ccy, 1.0)
            transfers.append({
                "Source Account": src["account"],
                "Source Account Venue Location": rnd.choice(uni.venues),
                "Target Account": tgt["account"],
                "Target Account Venue Location": rnd.choice(uni.venues),
                "Currency": ccy,
                "Reference": reference,
                "Value Date": day,
                "Value Amount": amount,
                "CCY (Display)": "GBP",
                "Value Amount (Display)": round(amount * rate, 3),
                "Sending Strategy": _weighted(rnd, strategy_mix),
                "Comment": rnd.choice(["Funding move", "Liquidity sweep",
                                       "Balance alignment", "Intraday top-up", ""]),
                "Created Time": created,
                "Created By": rnd.choice(uni.people),
                "Approved Time": approved_time,
                "Approved By": approved_by,
                "Transfer Status": status,
                "Message Status": _weighted(rnd, message_mix),
                "Actions": "View details",
            })

        # ---------------- business ledger ----------------
        for _ in range(int(rnd.gauss(420, 45) * factor)):
            pair = uni.pick_active()
            ccy = pair["currency"]
            rate = FX_TO_GBP.get(ccy, 1.0)
            mark = "DR" if rnd.random() < 0.5 else "CR"
            local = _amount_for(rnd, digit_mix) * (1 if mark == "CR" else -1)
            stamp = _timestamp(rnd, day)
            # Some postings trace back to a transfer, which is what makes the
            # views joinable.
            upstream = (transfers[-rnd.randint(1, min(40, len(transfers)))]["Reference"]
                        if transfers and rnd.random() < 0.45 else None)
            ledger.append({
                "Account": pair["account"],
                "Cashflow Type": rnd.choice(CASHFLOW_TYPES),
                "Sub Branch": pair["desk"],
                "Value Date": day,
                "Debit/Credit Mark": mark,
                "CCY (Local)": ccy,
                "Amount (Local)": local,
                "CCY (Display)": "GBP",
                "Amount (Display)": round(local * rate, 3),
                "Transaction Timestamp": stamp,
                "Transaction Reference": f"CF{day:%y%m%d} {txn_seq:05d}",
                "Ledger Account": f"GL.NOSTRO.{ccy}.{pair['desk']}.{rnd.randint(1, 9):03d}",
                "Counterparty": rnd.choice(uni.counterparties),
                "Book Id": pair["desk"] + rnd.choice(BOOK_SUFFIX),
                "Trade Status": rnd.choices(TRADE_STATUSES, weights=[0.86, 0.11, 0.03])[0],
                "Upstream Transaction ID": upstream,
                "Net Type": rnd.choices(NET_TYPES, weights=[0.55, 0.33, 0.12])[0],
            })
            txn_seq += 1

        # ---------------- client balances (one row per active pair per day) ----
        for pair in uni.active:
            ccy = pair["currency"]
            rate = FX_TO_GBP.get(ccy, 1.0)
            base = 10 ** rnd.uniform(4.6, 8.4) / max(rate, 0.001) ** 0.35
            sod = round(base * rnd.uniform(0.7, 1.3), 3)
            busy = rnd.random() < 0.62
            credits = round(base * rnd.uniform(0.02, 0.55), 3) if busy else 0.0
            debits = -round(base * rnd.uniform(0.02, 0.5), 3) if busy else 0.0
            calculated = round(sod + credits + debits, 3)

            # Reconciliation breaks: rare, and larger on the stressed day.
            if rnd.random() < (0.075 if day == spike_day else 0.022):
                drift = round(abs(calculated) * rnd.uniform(0.0004, 0.02), 3) or 12.5
                eod = round(calculated - drift, 3)
            else:
                eod = calculated

            last_txn = (_timestamp(rnd, day) if busy else None)
            client.append({
                "Legal Entity": pair["entity"],
                "Currency": ccy,
                "Account": pair["account"],
                "Sort Code": pair["sort_code"],
                "Account Name": pair["name"],
                "Value Date": day,
                "Start of Day Balance (Local)": sod,
                "Credits (Local)": credits,
                "Debits (Local)": debits,
                "Calculated Balance (Local)": calculated,
                "Swing (Local)": round(calculated - sod, 3),
                "Last Transaction Received": last_txn,
                "EOD Balance (Local)": eod,
                "Difference (Local)": round(calculated - eod, 3),
                "Match": "✓" if abs(calculated - eod) <= 0.01 else "✕",
                "Start of Day Balance (Display)": round(sod * rate, 3),
                "Credits (Display)": round(credits * rate, 3),
                "Debits (Display)": round(debits * rate, 3),
                "Calculated Balance (Display)": round(calculated * rate, 3),
                "Swing (Display)": round((calculated - sod) * rate, 3),
            })

    return {"transfers": transfers, "ledger": ledger, "client": client,
            "universe": uni, "dates": dates,
            "spike_day": spike_day, "quiet_day": quiet_day}


# --------------------------------------------------------------------------
# Workbook writing - matches the layout the application already expects:
# title, two control rows, a blank row, header on row 6, then a Total footer.
# --------------------------------------------------------------------------
TITLE_FONT = Font(bold=True, size=13)
HEAD_FONT = Font(bold=True, size=10)
HEAD_FILL = PatternFill("solid", fgColor="F2EAFB")
CTRL_FONT = Font(size=9, color="1F4E79")


def _write_view(ws, title: str, controls: list[tuple], rows: list[dict],
                columns: list[str], number_formats: dict[str, str]) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    for r, pairs in enumerate(controls, start=2):
        col = 1
        for key, value in pairs:
            ws.cell(row=r, column=col, value=key).font = CTRL_FONT
            ws.cell(row=r, column=col + 1, value=value).font = CTRL_FONT
            col += 3

    header_row = 6
    for j, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for i, row in enumerate(rows, start=header_row + 1):
        for j, name in enumerate(columns, start=1):
            value = row.get(name)
            cell = ws.cell(row=i, column=j, value=value)
            fmt = number_formats.get(name)
            if fmt:
                cell.number_format = fmt

    ws.cell(row=header_row + len(rows) + 1, column=1, value="Total").font = HEAD_FONT

    for j, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(13, len(name) + 4), 30)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_workbook(data: dict, out: Path, profile: dict, args) -> None:
    wb = Workbook()
    uni, dates = data["universe"], data["dates"]
    first, last = dates[0], dates[-1]
    money = "#,##0.000"
    date_fmt = "dd/mm/yyyy"
    stamp_fmt = "dd/mm/yyyy hh:mm"

    # README ---------------------------------------------------------------
    ws = wb.active
    ws.title = "README"
    lines = [
        ("Synthetic Liquidity Views - one month", True),
        ("Generated data. Every identifier, name, counterparty and amount is fabricated.", False),
        ("", False),
        (f"Period: {first:%d %b %Y} to {last:%d %b %Y} ({len(dates)} business days)", False),
        (f"Account/currency pairs in the universe: {len(uni.pairs):,}", False),
        (f"Pairs with activity in the period: {len(uni.active):,}", False),
        ("", False),
        ("Shape derived from an OCR profile of production screenshots:", True),
        (f"  currency mix over {len(profile.get('currency_mix', {}))} currencies", False),
        ("  transfer status and sending strategy mix", False),
        ("  transaction value magnitudes", False),
        ("Identifiers were NOT derived from that source - the OCR recovered none.", False),
        ("", False),
        ("Deliberate features, so reports have something to find:", True),
        (f"  elevated activity on {data['spike_day']:%d %b} (a stressed day)", False),
        (f"  reduced activity on {data['quiet_day']:%d %b} (a quiet day)", False),
        ("  reconciliation breaks on a small share of account-days", False),
        ("  failed and pending transfers, and rejected messages", False),
        ("  month-end and day-of-week volume effects", False),
        ("  an intraday arrival curve with morning and afternoon peaks", False),
        ("", False),
        ("Upstream Transaction ID on the ledger references a transfer Reference,", False),
        ("so the transfer and ledger views can be combined.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 96

    # Views ----------------------------------------------------------------
    transfer_cols = ["Source Account", "Source Account Venue Location", "Target Account",
                     "Target Account Venue Location", "Currency", "Reference", "Value Date",
                     "Value Amount", "CCY (Display)", "Value Amount (Display)",
                     "Sending Strategy", "Comment", "Created Time", "Created By",
                     "Approved Time", "Approved By", "Transfer Status", "Message Status",
                     "Actions"]
    _write_view(
        wb.create_sheet("Nostro Transfer View"), "Nostro Transfer View",
        [[("Group", "SYNTHETIC FUNDING ALL ENTITY"), ("Value Date", first),
          ("To", last), ("Time Zone", "London"), ("Display Currency", "GBP")],
         [("Scope", "All"), ("Items Loaded", len(data["transfers"])),
          ("Data Classification", "Synthetic / non-production")]],
        data["transfers"], transfer_cols,
        {"Value Amount": money, "Value Amount (Display)": money,
         "Value Date": date_fmt, "Created Time": stamp_fmt, "Approved Time": stamp_fmt})

    client_cols = ["Legal Entity", "Currency", "Account", "Sort Code", "Account Name",
                   "Value Date", "Start of Day Balance (Local)", "Credits (Local)",
                   "Debits (Local)", "Calculated Balance (Local)", "Swing (Local)",
                   "Last Transaction Received", "EOD Balance (Local)",
                   "Difference (Local)", "Match", "Start of Day Balance (Display)",
                   "Credits (Display)", "Debits (Display)",
                   "Calculated Balance (Display)", "Swing (Display)"]
    _write_view(
        wb.create_sheet("Client View"), "Client View",
        [[("Client Group", "SYNTHETIC CLIENT GROUP"), ("Value Date", first),
          ("To", last), ("Time Zone", "London"), ("Display Currency", "GBP"),
          ("Match Tolerance", 0.01)],
         [("Refresh Mode", "Manual"), ("Rows", len(data["client"])),
          ("Data Classification", "Synthetic / non-production")]],
        data["client"], client_cols,
        {c: money for c in client_cols if "Balance" in c or "Credits" in c
         or "Debits" in c or "Swing" in c or "Difference" in c}
        | {"Value Date": date_fmt, "Last Transaction Received": stamp_fmt})

    ledger_cols = ["Account", "Cashflow Type", "Sub Branch", "Value Date",
                   "Debit/Credit Mark", "CCY (Local)", "Amount (Local)", "CCY (Display)",
                   "Amount (Display)", "Transaction Timestamp", "Transaction Reference",
                   "Ledger Account", "Counterparty", "Book Id", "Trade Status",
                   "Upstream Transaction ID", "Net Type"]
    _write_view(
        wb.create_sheet("Business Ledger Txn View"), "Business Ledger Transaction View",
        [[("Business Group", "SYNTHETIC CASH MANAGEMENT"), ("Value Date From", first),
          ("Value Date To", last), ("Display CCY", "GBP")],
         [("Dr/Cr Mark", "Both"), ("Include Cancelled", "Include"),
          ("Rows", len(data["ledger"])),
          ("Data Classification", "Synthetic / non-production")]],
        data["ledger"], ledger_cols,
        {"Amount (Local)": money, "Amount (Display)": money,
         "Value Date": date_fmt, "Transaction Timestamp": stamp_fmt})

    _write_dictionary(wb.create_sheet("Data Dictionary"),
                      transfer_cols, client_cols, ledger_cols)
    _write_controls(wb.create_sheet("View Controls"), first, last, data)
    _write_reference(wb.create_sheet("Reference Data"))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def _write_dictionary(ws, transfer_cols, client_cols, ledger_cols) -> None:
    """The application reads its caveats from this sheet, so it must describe the
    generated columns - especially the derived ones."""
    ws["A1"] = "Data Dictionary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Column-level attributes for the three synthetic views"
    header = ["View", "Ordinal", "Column Name", "Data Type", "Length / Precision",
              "Excel Format", "Nullable", "Derived", "Allowed Values / Domain",
              "Definition / Business Rule", "Formula / Calculation",
              "Synthetic Example", "Screenshot Interpretation"]
    for j, name in enumerate(header, start=1):
        cell = ws.cell(row=4, column=j, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL

    derived_rules = {
        "Calculated Balance (Local)": "=Start of Day Balance + Credits + Debits",
        "Swing (Local)": "=Calculated Balance - Start of Day Balance",
        "Difference (Local)": "=Calculated Balance - EOD Balance",
        "Match": '=IF(ABS(Difference)<=Tolerance,"tick","cross")',
        "Value Amount (Display)": "=Value Amount * FX rate to display currency",
        "Amount (Display)": "=Amount (Local) * FX rate to display currency",
        "Start of Day Balance (Display)": "=Start of Day Balance (Local) * FX rate",
        "Credits (Display)": "=Credits (Local) * FX rate",
        "Debits (Display)": "=Debits (Local) * FX rate",
        "Calculated Balance (Display)": "=Calculated Balance (Local) * FX rate",
        "Swing (Display)": "=Swing (Local) * FX rate",
    }
    domains = {
        "Currency": "ISO 4217 currency code",
        "CCY (Local)": "ISO 4217 currency code",
        "CCY (Display)": "GBP",
        "Sending Strategy": ", ".join(sorted(DOMAINS_SEEN.get("strategy", []))),
        "Transfer Status": ", ".join(["LEDGER_OR_CASHFLOW_RECEIVED", "LEDGER_RECEIVED",
                                      "CASHFLOW_RECEIVED", "PENDING_APPROVAL", "FAILED"]),
        "Message Status": ", ".join(sorted(DOMAINS_SEEN.get("message", []))),
        "Debit/Credit Mark": "DR, CR",
        "Trade Status": "NEW, AMEND, CANCEL",
        "Net Type": "NETTER, GROSS, INTERNAL",
        "Match": "tick or cross",
        "Upstream Transaction ID": "References Reference on the Nostro Transfer View",
    }
    row = 5
    for view_name, cols in [("Nostro Transfer View", transfer_cols),
                            ("Client View", client_cols),
                            ("Business Ledger Transaction View", ledger_cols)]:
        for i, name in enumerate(cols, start=1):
            derived = name in derived_rules
            money = any(k in name for k in ("Amount", "Balance", "Credits", "Debits",
                                            "Swing", "Difference"))
            ws.cell(row=row, column=1, value=view_name)
            ws.cell(row=row, column=2, value=i)
            ws.cell(row=row, column=3, value=name)
            ws.cell(row=row, column=4, value="Decimal" if money else
                    ("DateTime" if "Time" in name else
                     ("Date" if "Date" in name else "Text")))
            ws.cell(row=row, column=6, value="#,##0.000" if money else "Text")
            ws.cell(row=row, column=7, value="Yes" if name in
                    ("Approved Time", "Approved By", "Comment", "Upstream Transaction ID",
                     "Last Transaction Received") else "No")
            ws.cell(row=row, column=8, value="Yes" if derived else "No")
            ws.cell(row=row, column=9, value=domains.get(name, ""))
            ws.cell(row=row, column=10,
                    value=f"{name} on the {view_name}.")
            ws.cell(row=row, column=11, value=derived_rules.get(name, ""))
            ws.cell(row=row, column=13, value="Generated")
            row += 1
    for j, width in enumerate([30, 8, 32, 12, 16, 12, 10, 9, 46, 46, 44, 20, 22], start=1):
        ws.column_dimensions[get_column_letter(j)].width = width


def _write_controls(ws, first, last, data) -> None:
    ws["A1"] = "View Controls and Filters"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Selectors in effect for the generated period"
    for j, name in enumerate(["View", "Control / Filter", "Data Type",
                              "Synthetic Value", "Definition"], start=1):
        cell = ws.cell(row=4, column=j, value=name)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL

    rows = [
        ("Nostro Transfer View", "Group", "Text", "SYNTHETIC FUNDING ALL ENTITY",
         "Funding group selector."),
        ("Nostro Transfer View", "Value Date From", "Date", f"{first:%Y-%m-%d}",
         "Start of the value-date range."),
        ("Nostro Transfer View", "Value Date To", "Date", f"{last:%Y-%m-%d}",
         "End of the value-date range."),
        ("Nostro Transfer View", "Display Currency", "ISO 4217", "GBP",
         "Currency used for the display amount."),
        ("Nostro Transfer View", "Time Zone", "Enum", "London", "Display time zone."),
        ("Client View", "Client Group", "Text", "SYNTHETIC CLIENT GROUP",
         "Client group selector."),
        ("Client View", "Value Date From", "Date", f"{first:%Y-%m-%d}",
         "Start of the snapshot range."),
        ("Client View", "Value Date To", "Date", f"{last:%Y-%m-%d}",
         "End of the snapshot range."),
        ("Client View", "Match Tolerance", "Decimal", "0.01",
         "Absolute difference below which a balance counts as matched."),
        ("Client View", "Display Currency", "ISO 4217", "GBP",
         "Currency used for display balances."),
        ("Business Ledger Transaction View", "Business Group", "Text",
         "SYNTHETIC CASH MANAGEMENT", "Business group selector."),
        ("Business Ledger Transaction View", "Value Date From", "Date",
         f"{first:%Y-%m-%d}", "Start of value-date range."),
        ("Business Ledger Transaction View", "Value Date To", "Date",
         f"{last:%Y-%m-%d}", "End of value-date range."),
        ("Business Ledger Transaction View", "Include Cancelled", "Enum", "Include",
         "Cancelled trades are included in this generated set."),
        ("Business Ledger Transaction View", "Display CCY", "ISO 4217", "GBP",
         "Currency used for the display amount."),
    ]
    for i, values in enumerate(rows, start=5):
        for j, value in enumerate(values, start=1):
            ws.cell(row=i, column=j, value=value)
    for j, width in enumerate([34, 22, 12, 32, 60], start=1):
        ws.column_dimensions[get_column_letter(j)].width = width


def _write_reference(ws) -> None:
    ws["A1"] = "Reference Data - synthetic support values"
    ws["A1"].font = TITLE_FONT
    ws.cell(row=3, column=1, value="Currency").font = HEAD_FONT
    ws.cell(row=3, column=2, value="GBP per 1 local unit").font = HEAD_FONT
    for i, (ccy, rate) in enumerate(sorted(FX_TO_GBP.items()), start=4):
        ws.cell(row=i, column=1, value=ccy)
        ws.cell(row=i, column=2, value=rate)
    ws.cell(row=len(FX_TO_GBP) + 5, column=1, value="Important")
    ws.cell(row=len(FX_TO_GBP) + 5, column=2,
            value="All FX rates are synthetic, flat across the period, and included "
                  "only to support display-currency translation.")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 88


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--profile", type=Path,
                        default=Path("anonymisation") / "ocr_profile.json")
    parser.add_argument("--out", type=Path,
                        default=Path("data") / "synthetic_liquidity_month.xlsx")
    parser.add_argument("--accounts", type=int, default=6000,
                        help="account/currency pairs in the universe")
    parser.add_argument("--active", type=int, default=900,
                        help="pairs with activity during the period")
    parser.add_argument("--days", type=int, default=22, help="business days")
    parser.add_argument("--end", type=str, default=None, help="last business day, YYYY-MM-DD")
    parser.add_argument("--seed", type=int, default=20260819)
    args = parser.parse_args()

    profile = json.loads(args.profile.read_text(encoding="utf-8")) if args.profile.exists() else {}
    if not profile:
        print(f"No profile at {args.profile}; falling back to built-in defaults.")

    end = (dt.date.fromisoformat(args.end) if args.end else dt.date.today())
    while end.weekday() >= 5:
        end -= dt.timedelta(days=1)

    print(f"Generating {args.days} business days ending {end:%d %b %Y}...")
    data = generate(profile, args.days, args.accounts, args.active, end, args.seed)
    write_workbook(data, args.out, profile, args)

    size_mb = args.out.stat().st_size / 1_048_576
    print(f"\nWritten -> {args.out}  ({size_mb:.1f} MB)")
    print(f"  Nostro transfers      {len(data['transfers']):>7,}")
    print(f"  Business ledger txns  {len(data['ledger']):>7,}")
    print(f"  Client balance rows   {len(data['client']):>7,}")
    print(f"  Account/ccy universe  {len(data['universe'].pairs):>7,}")
    print(f"  Active in period      {len(data['universe'].active):>7,}")
    print(f"  Currencies            {len(data['universe'].currency_mix):>7,}")
    print(f"\n  Stressed day: {data['spike_day']:%a %d %b}   "
          f"Quiet day: {data['quiet_day']:%a %d %b}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
