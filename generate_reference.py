"""Build the reference tables described in reference_model.py.

Keys are taken from the values that actually appear in the live data, so every
row joins. Attributes are generated: invented names, but real countries,
regions, ISO codes and currencies, because a jurisdiction question answered with
invented geography tells you nothing.

Writes data/reference_data.xlsx and DATA_MODEL.md from the same model, so the
documentation always describes the tables that exist.

Usage:
    python generate_reference.py
"""

from __future__ import annotations

import argparse
import random
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import reference_model as model
from app import data_access

# --------------------------------------------------------------------------
# Real-world reference. Countries, regions and currencies are genuine; the
# risk and sanctions columns are illustrative structure only.
# --------------------------------------------------------------------------
COUNTRIES = {
    # country: (ISO, region, sub region, sanctions regime, FATF, EU/EEA, risk tier, ccy)
    "United Kingdom": ("GB", "Europe", "Northern Europe", "None", "None", "No", "Low", "GBP"),
    "United States": ("US", "Americas", "Northern America", "None", "None", "No", "Low", "USD"),
    "Germany": ("DE", "Europe", "Western Europe", "None", "None", "Yes", "Low", "EUR"),
    "France": ("FR", "Europe", "Western Europe", "None", "None", "Yes", "Low", "EUR"),
    "Netherlands": ("NL", "Europe", "Western Europe", "None", "None", "Yes", "Low", "EUR"),
    "Ireland": ("IE", "Europe", "Northern Europe", "None", "None", "Yes", "Low", "EUR"),
    "Luxembourg": ("LU", "Europe", "Western Europe", "None", "None", "Yes", "Low", "EUR"),
    "Switzerland": ("CH", "Europe", "Western Europe", "None", "None", "No", "Low", "CHF"),
    "Spain": ("ES", "Europe", "Southern Europe", "None", "None", "Yes", "Low", "EUR"),
    "Italy": ("IT", "Europe", "Southern Europe", "None", "None", "Yes", "Low", "EUR"),
    "Poland": ("PL", "Europe", "Eastern Europe", "None", "None", "Yes", "Low", "PLN"),
    "Czechia": ("CZ", "Europe", "Eastern Europe", "None", "None", "Yes", "Low", "CZK"),
    "Hungary": ("HU", "Europe", "Eastern Europe", "None", "None", "Yes", "Medium", "HUF"),
    "Sweden": ("SE", "Europe", "Northern Europe", "None", "None", "Yes", "Low", "SEK"),
    "Norway": ("NO", "Europe", "Northern Europe", "None", "None", "No", "Low", "NOK"),
    "Denmark": ("DK", "Europe", "Northern Europe", "None", "None", "Yes", "Low", "DKK"),
    "Japan": ("JP", "Asia", "Eastern Asia", "None", "None", "No", "Low", "JPY"),
    "Singapore": ("SG", "Asia", "South-eastern Asia", "None", "None", "No", "Low", "SGD"),
    "Hong Kong SAR": ("HK", "Asia", "Eastern Asia", "None", "None", "No", "Medium", "HKD"),
    "China": ("CN", "Asia", "Eastern Asia", "Targeted", "None", "No", "Medium", "CNY"),
    "India": ("IN", "Asia", "Southern Asia", "None", "None", "No", "Medium", "INR"),
    "South Korea": ("KR", "Asia", "Eastern Asia", "None", "None", "No", "Low", "KRW"),
    "Taiwan": ("TW", "Asia", "Eastern Asia", "None", "None", "No", "Medium", "TWD"),
    "Thailand": ("TH", "Asia", "South-eastern Asia", "None", "None", "No", "Medium", "THB"),
    "Australia": ("AU", "Oceania", "Australia and NZ", "None", "None", "No", "Low", "AUD"),
    "New Zealand": ("NZ", "Oceania", "Australia and NZ", "None", "None", "No", "Low", "NZD"),
    "Canada": ("CA", "Americas", "Northern America", "None", "None", "No", "Low", "CAD"),
    "Mexico": ("MX", "Americas", "Latin America", "None", "None", "No", "Medium", "MXN"),
    "Brazil": ("BR", "Americas", "Latin America", "None", "None", "No", "Medium", "BRL"),
    "South Africa": ("ZA", "Africa", "Sub-Saharan Africa", "None", "Grey list", "No", "Medium", "ZAR"),
    "United Arab Emirates": ("AE", "Asia", "Western Asia", "None", "None", "No", "Medium", "AED"),
    "Saudi Arabia": ("SA", "Asia", "Western Asia", "None", "None", "No", "Medium", "SAR"),
    "Turkey": ("TR", "Asia", "Western Asia", "None", "None", "No", "High", "TRY"),
    "Israel": ("IL", "Asia", "Western Asia", "None", "None", "No", "Medium", "ILS"),
    "Romania": ("RO", "Europe", "Eastern Europe", "None", "None", "Yes", "Medium", "RON"),
    # The jurisdictions that make the question worth asking.
    "Russia": ("RU", "Europe", "Eastern Europe", "Comprehensive", "Suspended", "No", "Prohibited", "RUB"),
    "Belarus": ("BY", "Europe", "Eastern Europe", "Comprehensive", "None", "No", "Prohibited", "BYN"),
    "Iran": ("IR", "Asia", "Southern Asia", "Comprehensive", "Black list", "No", "Prohibited", "IRR"),
    "Syria": ("SY", "Asia", "Western Asia", "Comprehensive", "None", "No", "Prohibited", "SYP"),
    "Myanmar": ("MM", "Asia", "South-eastern Asia", "Sectoral", "Black list", "No", "High", "MMK"),
    "Nigeria": ("NG", "Africa", "Sub-Saharan Africa", "None", "Grey list", "No", "High", "NGN"),
    "Panama": ("PA", "Americas", "Latin America", "None", "Grey list", "No", "High", "PAB"),
    "Cayman Islands": ("KY", "Americas", "Caribbean", "None", "None", "No", "High", "KYD"),
}

CITY_COUNTRY = {
    "London": "United Kingdom", "New York": "United States", "Frankfurt": "Germany",
    "Zurich": "Switzerland", "Singapore": "Singapore", "Tokyo": "Japan",
    "Hong Kong": "Hong Kong SAR", "Sydney": "Australia", "Luxembourg": "Luxembourg",
    "Toronto": "Canada", "Mexico City": "Mexico", "Warsaw": "Poland",
    "Stockholm": "Sweden", "Oslo": "Norway", "Dublin": "Ireland",
    "Paris": "France", "Madrid": "Spain", "Milan": "Italy",
}
CITY_TZ = {
    "London": "Europe/London", "New York": "America/New_York",
    "Frankfurt": "Europe/Berlin", "Zurich": "Europe/Zurich",
    "Singapore": "Asia/Singapore", "Tokyo": "Asia/Tokyo",
    "Hong Kong": "Asia/Hong_Kong", "Sydney": "Australia/Sydney",
    "Luxembourg": "Europe/Luxembourg", "Toronto": "America/Toronto",
    "Mexico City": "America/Mexico_City", "Warsaw": "Europe/Warsaw",
    "Stockholm": "Europe/Stockholm", "Oslo": "Europe/Oslo",
    "Dublin": "Europe/Dublin", "Paris": "Europe/Paris",
    "Madrid": "Europe/Madrid", "Milan": "Europe/Rome",
}
CCY_NAMES = {
    "GBP": "Pound Sterling", "USD": "US Dollar", "EUR": "Euro", "JPY": "Japanese Yen",
    "CHF": "Swiss Franc", "HKD": "Hong Kong Dollar", "SGD": "Singapore Dollar",
    "AUD": "Australian Dollar", "NOK": "Norwegian Krone", "SEK": "Swedish Krona",
    "ZAR": "South African Rand", "CNY": "Chinese Yuan", "INR": "Indian Rupee",
    "CAD": "Canadian Dollar", "NZD": "New Zealand Dollar", "DKK": "Danish Krone",
    "PLN": "Polish Zloty", "MXN": "Mexican Peso", "TRY": "Turkish Lira",
    "AED": "UAE Dirham", "SAR": "Saudi Riyal", "THB": "Thai Baht",
    "KRW": "South Korean Won", "TWD": "Taiwan Dollar", "BRL": "Brazilian Real",
    "CZK": "Czech Koruna", "HUF": "Hungarian Forint", "ILS": "Israeli Shekel",
    "RON": "Romanian Leu",
}
CLS_CURRENCIES = {"GBP", "USD", "EUR", "JPY", "CHF", "AUD", "CAD", "DKK", "NOK",
                  "SEK", "SGD", "HKD", "NZD", "ZAR", "KRW", "ILS", "MXN", "HUF"}
RESTRICTED_CURRENCIES = {"CNY", "INR", "BRL", "TWD", "KRW", "RUB"}

SECTORS = ["Banking", "Broker Dealer", "Insurance", "Asset Management",
           "Central Banking", "Market Infrastructure", "Corporate Treasury"]
FIRST = ["Avery", "Blake", "Casey", "Devon", "Emery", "Finley", "Greer", "Harper",
         "Jules", "Kai", "Lennox", "Marlow", "Noor", "Oakley", "Quinn", "Reese"]
LAST = ["Ashford", "Brookes", "Calloway", "Delaney", "Ellison", "Fairweather",
        "Hollis", "Jennings", "Kirby", "Merrick", "Novak", "Prentice", "Rhodes"]
LEGAL_SUFFIX = ["plc", "AG", "S.A.", "N.V.", "Ltd", "Inc.", "Pte Ltd", "GmbH"]


class Generator:
    """Deterministic per-key generation, so a value never changes between runs."""

    def __init__(self, seed: int) -> None:
        self.seed = seed
        self.fk_pools: dict[str, list[str]] = {}

    def rnd(self, table: str, key: str, attribute: str) -> random.Random:
        return random.Random(f"{self.seed}:{table}:{key}:{attribute}")

    def value(self, kind: str, table: str, key: str, attribute: str):
        r = self.rnd(table, key, attribute)

        if kind.startswith("choice:"):
            return r.choice(kind.split(":", 1)[1].split(","))

        if kind.startswith("weighted:"):
            pairs = [p.rsplit(":", 1) for p in kind.split(":", 1)[1].split(",")]
            options = [p[0] for p in pairs]
            weights = [float(p[1]) for p in pairs]
            return r.choices(options, weights=weights, k=1)[0]

        if kind.startswith("amount:"):
            low, high = (float(x) for x in kind.split(":")[1:3])
            return round(r.uniform(low, high), -3)

        if kind.startswith("fk:"):
            pool = self.fk_pools.get(kind[3:], [])
            return r.choice(pool) if pool else ""

        if kind == "date_past":
            return f"{r.randint(2008, 2022)}-{r.randint(1, 12):02d}-{r.randint(1, 28):02d}"
        if kind == "date_recent":
            return f"{r.choice([2025, 2026])}-{r.randint(1, 12):02d}-{r.randint(1, 28):02d}"
        if kind == "date_future":
            return f"{r.choice([2026, 2027])}-{r.randint(1, 12):02d}-{r.randint(1, 28):02d}"

        if kind == "lei":
            body = "".join(r.choice("ABCDEFGHJKLMNPQRSTUVWXYZ0123456789") for _ in range(18))
            return f"{body}{r.randint(10, 99)}"
        if kind == "bic":
            return ("".join(r.choice("ABCDEFGHJKLMNPQRSTUVWXYZ") for _ in range(4))
                    + "".join(r.choice("ABCDEFGHJKLMNPQRSTUVWXYZ") for _ in range(2))
                    + "".join(r.choice("ABCDEFGHJKLMNPQRSTUVWXYZ0123456789") for _ in range(2)))
        if kind == "iban":
            return ("GB" + f"{r.randint(10, 99)}" + "".join(
                r.choice("ABCDEFGHJKLMNPQRSTUVWXYZ") for _ in range(4))
                + "".join(str(r.randint(0, 9)) for _ in range(14)))
        if kind == "legal_name":
            return f"{key.split()[0] if key else 'Alder'} Holdings {r.choice(LEGAL_SUFFIX)}"
        if kind == "person_name":
            return f"{r.choice(FIRST)} {r.choice(LAST)}"
        if kind == "city":
            return r.choice(list(CITY_COUNTRY))
        if kind == "cost_centre":
            return f"CC{r.randint(1000, 9999)}"
        if kind == "user_id":
            initials = "".join(part[0] for part in key.split()[:2]).lower()
            return f"{initials}{r.randint(10000, 99999)}"

        # Values derived from the key itself.
        if kind == "city_from_name":
            return next((c for c in CITY_COUNTRY if c in key), "London")
        if kind == "timezone_from_city":
            city = next((c for c in CITY_TZ if c in key), None)
            return CITY_TZ.get(city or r.choice(list(CITY_TZ)), "Europe/London")
        if kind == "ccy_name":
            return CCY_NAMES.get(key, key)
        if kind == "ccy_convertibility":
            return "Restricted" if key in RESTRICTED_CURRENCIES else "Freely convertible"
        if kind == "ccy_cls":
            return "Yes" if key in CLS_CURRENCIES else "No"
        if kind == "ccy_restricted":
            return "Yes" if key in RESTRICTED_CURRENCIES else "No"
        if kind == "ccy_of_country":
            return COUNTRIES.get(key, ("", "", "", "", "", "", "", "GBP"))[7]
        if kind == "book_desk":
            return key[:4]
        if kind == "book_name":
            suffix = key[4:] or "MAIN"
            return f"{key[:4]} {suffix.title()} Book"
        if kind == "desk_name":
            return f"{key} Treasury Desk"
        if kind == "branch_name":
            return f"{key} Branch"
        if kind == "product_name":
            return f"{key} Product"
        if kind == "sector_desc":
            return f"{key} sector counterparties"
        if kind == "nace":
            return f"K{r.randint(64, 66)}.{r.randint(10, 99)}"
        if kind == "gl_desc":
            parts = key.split(".")
            return f"{parts[1].title()} account {parts[2]} at {parts[3]}" \
                if len(parts) > 3 else key
        if kind == "cashflow_desc":
            return f"Cashflow class {key}"
        if kind == "clearing_system":
            return r.choice(["CHAPS", "TARGET2", "Fedwire", "CHIPS", "RTGS", "SIC"])
        if kind == "operator_oversight":
            return r.choice(["Central Bank", "National Regulator", "Consortium Board"])
        if kind == "scheme_name":
            return {"SWIFT": "SWIFT Cross-border", "INTERNAL": "Internal Book Transfer"}.get(
                key, f"{key} Funding Settlement Service")
        if kind == "scheme_network":
            return {"SWIFT": "Correspondent network", "INTERNAL": "Intragroup"}.get(
                key, "Proprietary")
        if kind == "scheme_settlement":
            return {"SWIFT": "Deferred net", "INTERNAL": "On-us"}.get(key, "RTGS")
        if kind == "scheme_standard":
            return {"SWIFT": "ISO 15022 / MT", "INTERNAL": "Proprietary"}.get(key, "ISO 20022")

        # Country attributes, straight from the real table.
        if kind.startswith("country_"):
            row = COUNTRIES.get(key)
            if not row:
                return ""
            index = {"country_iso": 0, "country_region": 1, "country_subregion": 2,
                     "country_sanctions": 3, "country_fatf": 4, "country_eu": 5,
                     "country_risk": 6, "country_ccy": 7}[kind]
            return row[index]

        return ""


def live_keys() -> dict[str, list[str]]:
    """Distinct values from every transaction workbook present.

    A data lake covers the whole estate, not one extract, so keys are taken from
    all of them. Without this a reference table generated against one workbook
    joins to nothing when the app is pointed at another.
    """
    frames: dict[str, list[pd.DataFrame]] = {v: [] for v in data_access.VIEW_SHEETS}
    workbooks = sorted(Path("data").glob("synthetic_liquidity*.xlsx"))
    for workbook in workbooks:
        original, data_access.DATA_FILE = data_access.DATA_FILE, workbook
        data_access._workbook_frames.cache_clear()
        try:
            for view in data_access.VIEW_SHEETS:
                frames[view].append(data_access.get_frame(view))
        except Exception:  # a workbook that will not load is simply skipped
            pass
        finally:
            data_access.DATA_FILE = original
            data_access._workbook_frames.cache_clear()
    print(f"  keys taken from: {', '.join(w.name for w in workbooks)}")

    def distinct(pairs):
        out = set()
        for view, col in pairs:
            for df in frames[view]:
                if col in df.columns:
                    out |= {str(v).strip() for v in df[col].dropna() if str(v).strip()}
        return sorted(out)

    keys = {name: distinct(spec["joins_from"])
            for name, spec in model.ROUND_1.items()}
    keys["country_master"] = sorted(COUNTRIES)
    return keys


def build(seed: int) -> dict[str, pd.DataFrame]:
    gen = Generator(seed)
    keys = live_keys()
    tables: dict[str, pd.DataFrame] = {}

    # Round 2 pools that round 1 points at must exist before round 1 is built.
    gen.fk_pools["country_master"] = sorted(COUNTRIES)
    gen.fk_pools["industry_master"] = SECTORS
    gen.fk_pools["business_line_master"] = [
        "Payments and Cash Management", "Securities Services", "Markets Treasury",
        "Corporate Lending", "Financial Institutions Group"]
    gen.fk_pools["product_master"] = [f"PRD{n:03d}" for n in range(1, 13)]
    gen.fk_pools["branch_master"] = [f"BR{c[:3].upper()}" for c in list(CITY_COUNTRY)[:12]]
    gen.fk_pools["regulator_master"] = [
        "Prudential Regulation Authority", "Federal Reserve", "European Central Bank",
        "Monetary Authority of Singapore", "Financial Services Agency",
        "Swiss Financial Market Supervisory Authority"]
    gen.fk_pools["scheme_operator_master"] = [
        "Bank of England", "SWIFT SC", "Internal Treasury Platform", "EBA Clearing"]
    gen.fk_pools["group_master"] = sorted(
        {f"{n.split()[0]} Group" for n in keys.get("counterparty_master", [])}) or ["Alder Group"]
    gen.fk_pools["legal_entity_master"] = keys.get("legal_entity_master", [])
    gen.fk_pools["desk_master"] = keys.get("desk_master", [])

    # Purely random country assignment left Russia with no counterparties at
    # all, so the jurisdiction question this exists to demonstrate returned
    # nothing. A demo dataset has to contain the case being demonstrated.
    GUARANTEED_JURISDICTIONS = {
        "Russia": 4, "Belarus": 2, "Iran": 2, "Syria": 1, "Myanmar": 2,
        "Turkey": 3, "China": 3, "Cayman Islands": 2, "Panama": 1, "Nigeria": 1,
    }

    def make(name: str, spec: dict, key_values: list[str]) -> pd.DataFrame:
        key_col = spec["key"][0]
        rows = []
        for key in key_values:
            row = {key_col: key}
            for attr, kind, _note in spec["attributes"]:
                row[attr] = gen.value(kind, name, key, attr)
            rows.append(row)
        return pd.DataFrame(rows)

    for name, spec in model.ROUND_2.items():
        pool = gen.fk_pools.get(name) or keys.get(name) or []
        tables[name] = make(name, spec, pool)
    for name, spec in model.ROUND_1.items():
        tables[name] = make(name, spec, keys.get(name, []))

    # Place the guaranteed jurisdictions deterministically, taking the
    # counterparties in name order so a rerun produces the same allocation.
    cp = tables.get("counterparty_master")
    if cp is not None and "Country of Incorporation" in cp.columns:
        cp = cp.sort_values(cp.columns[0]).reset_index(drop=True)
        cursor = 0
        for country, wanted in GUARANTEED_JURISDICTIONS.items():
            for _ in range(wanted):
                if cursor >= len(cp):
                    break
                cp.loc[cursor, "Country of Incorporation"] = country
                cursor += 1
        tables["counterparty_master"] = cp
    return tables


def write_workbook(tables: dict[str, pd.DataFrame], out: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, size=10)
    head_fill = PatternFill("solid", fgColor="F2EAFB")

    for name, df in tables.items():
        ws = wb.create_sheet(name[:31])
        spec = model.ALL_TABLES[name]
        ws["A1"] = f"{name}  -  {spec['domain']} domain  -  round {model.round_of(name)}"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"Grain: {spec['grain']}. Generated reference data - names invented, " \
                   "countries and currencies real."
        ws["A2"].font = Font(size=9, italic=True, color="52514E")

        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=4, column=j, value=col)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for i, (_, row) in enumerate(df.iterrows(), start=5):
            for j, col in enumerate(df.columns, start=1):
                ws.cell(row=i, column=j, value=row[col])
        for j, col in enumerate(df.columns, start=1):
            width = max([len(str(col))] + [len(str(v)) for v in df[col].head(40)])
            ws.column_dimensions[get_column_letter(j)].width = min(max(14, width + 3), 34)
        ws.freeze_panes = ws.cell(row=5, column=1)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def write_documentation(tables: dict[str, pd.DataFrame], out: Path) -> None:
    lines = [
        "# Reference data model",
        "",
        "What a bank data lake would let us attach to this transaction data, and "
        "what each hop makes answerable.",
        "",
        "Generated from `reference_model.py` by `generate_reference.py`, so this "
        "document always describes the tables that actually exist. The data is "
        "synthetic: **names are invented, countries, regions, ISO codes and "
        "currencies are real**, because a jurisdiction question answered with "
        "invented geography tells you nothing.",
        "",
        "> The sanctions, FATF and risk-tier columns are illustrative structure so "
        "the shape of such a question can be demonstrated. **They are not a "
        "compliance source and must never be used for screening.**",
        "",
        "## Why two rounds",
        "",
        "Round 1 attaches to a column that exists in the live views. Round 2 "
        "attaches to something round 1 produced, so it is only reachable in two "
        "hops - and that is where the interesting questions live. \"Intraday usage "
        "for Russian counterparties\" is not answerable from the transaction data "
        "or from round 1 alone; it needs counterparty → country of incorporation → "
        "sanctions regime.",
        "",
    ]

    for round_no, group in ((1, model.ROUND_1), (2, model.ROUND_2)):
        lines += [f"## Round {round_no}", ""]
        if round_no == 1:
            lines += ["Keyed directly off a column in the live data.", ""]
        else:
            lines += ["Keyed off an attribute produced by round 1.", ""]

        for name, spec in group.items():
            df = tables.get(name, pd.DataFrame())
            lines += [f"### `{name}` — {spec['domain']}", ""]
            if round_no == 1:
                joins = ", ".join(f"`{v}.{c}`" for v, c in spec["joins_from"])
                lines.append(f"**Joins from:** {joins}")
            else:
                lines.append("**Reached via:** "
                             + ", ".join(f"`{v}`" for v in spec["reached_via"]))
            lines += [f"  ", f"**Grain:** {spec['grain']}  ",
                      f"**Rows generated:** {len(df):,}", "",
                      "| Attribute | Notes |", "|---|---|"]
            for attr, kind, note in spec["attributes"]:
                detail = note
                if kind.startswith("fk:"):
                    detail = (detail + " " if detail else "") + f"→ `{kind[3:]}`"
                lines.append(f"| {attr} | {detail} |")
            lines.append("")

    lines += [
        "## What this makes answerable",
        "",
        "| Question | Path |",
        "|---|---|",
        "| Intraday usage for Russian counterparties | ledger → `counterparty_master` "
        "→ `country_master`.Sanctions Regime |",
        "| Exposure by jurisdiction risk tier | ledger → `counterparty_master` → "
        "`country_master`.Jurisdiction Risk Tier |",
        "| Flow through restricted currencies | ledger → `currency_master`"
        ".Restricted Currency |",
        "| Usage by regulated entity and its regulator | client → "
        "`legal_entity_master` → `regulator_master` |",
        "| Which desks carry the systemic counterparties | ledger → "
        "`counterparty_master` → `group_master`.Globally Systemic |",
        "| Settlement concentration by scheme operator | transfers → "
        "`payment_scheme_master` → `scheme_operator_master` |",
        "| Dormant accounts still holding balance | client → `account_master`"
        ".Account Status |",
        "| Value at venues with a late cut-off | transfers → `venue_master`"
        ".Cut-off Time |",
        "",
        "## What is still not answerable",
        "",
        "Reference data adds attributes to things that already appear. It cannot "
        "create records that were never captured:",
        "",
        "- **why a payment failed** — no reason is recorded on the transaction",
        "- **the dormant account estate** — accounts with no activity are absent "
        "from the extract entirely, so no lookup can size them",
        "- **legal entity on a ledger posting** — the ledger carries no entity key, "
        "so `legal_entity_master` cannot be reached from it",
        "- **the transfer that produced a posting**, for the 55% of ledger rows "
        "with no upstream reference",
        "",
    ]
    out.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=Path("data") / "reference_data.xlsx")
    parser.add_argument("--docs", type=Path, default=Path("DATA_MODEL.md"))
    parser.add_argument("--seed", type=int, default=20260820)
    args = parser.parse_args()

    tables = build(args.seed)
    write_workbook(tables, args.out)
    write_documentation(tables, args.docs)

    total = sum(len(df) for df in tables.values())
    print(f"Wrote {len(tables)} reference tables ({total:,} rows) -> {args.out}")
    print(f"Documentation -> {args.docs}\n")
    for name, df in sorted(tables.items(), key=lambda kv: model.round_of(kv[0])):
        print(f"  round {model.round_of(name)}  {name:<26} {len(df):>5} rows  "
              f"{len(df.columns)} cols")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
