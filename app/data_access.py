"""Deterministic data access over the synthetic liquidity workbook.

Every figure that reaches a report comes from this module. The LLM chooses
*which* query to run (view, filters, grouping, measure); it never supplies or
computes the numbers themselves. That keeps reconciliation and ledger figures
auditable and reproducible.
"""

from __future__ import annotations

import datetime as dt
import os
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# The month-long generated set is the dataset: it carries the intraday shape,
# cardinality and joinable keys the tool needs, and it is the one that has been
# through the anonymiser. The original single-day sample it replaced held values
# the golden source marks for replacement, so it was removed rather than kept as
# a fallback. Override with DATA_FILE to pin a specific workbook.
DATA_FILE = Path(os.getenv("DATA_FILE")
                 or DATA_DIR / "synthetic_liquidity_month.xlsx")

# All three view sheets share the same layout: title row, two control rows,
# two blank rows, then the header on row 6. The final row is a "Total" footer
# that must not be treated as data.
HEADER_ROW = 6
TOTAL_ROW_MARKER = "Total"

VIEW_SHEETS = {
    "nostro_transfer": "Nostro Transfer View",
    "client": "Client View",
    "business_ledger": "Business Ledger Txn View",
}

# The Data Dictionary and View Controls sheets label the third view
# "Business Ledger Transaction View" while its sheet tab is abbreviated to
# "Txn". Metadata lookups must use these labels, not the sheet names.
DICTIONARY_LABELS = {
    "nostro_transfer": "Nostro Transfer View",
    "client": "Client View",
    "business_ledger": "Business Ledger Transaction View",
}

# Those two sheets carry a title, a subtitle and a blank row before the header.
METADATA_HEADER_ROW = 4

# Some columns hold glyphs (the reconciliation Match tick/cross). A language
# model cannot reliably reproduce these characters, so plain-word aliases are
# accepted for them and normalised back to the stored glyph before filtering.
GLYPH_ALIASES = {
    "✓": ["matched", "match", "true", "yes", "y", "pass", "passed", "ok", "tick", "check"],
    "✕": ["unmatched", "not matched", "no match", "false", "no", "n", "fail",
               "failed", "mismatch", "mismatched", "break", "cross", "x"],
}
_ALIAS_LOOKUP = {alias: glyph for glyph, aliases in GLYPH_ALIASES.items() for alias in aliases}

# Which column states the currency of each row, per view. Amounts denominated in
# that currency cannot be added together across different values of it.
CURRENCY_COLUMN = {
    "nostro_transfer": "Currency",
    "client": "Currency",
    "business_ledger": "CCY (Local)",
}

# Local-currency amount columns and their FX-translated equivalent. Aggregating
# a local column across more than one currency is adding unlike units; where a
# display equivalent exists the query is switched to it, and where none exists
# the query is refused.
DISPLAY_EQUIVALENT = {
    "nostro_transfer": {
        "Value Amount": "Value Amount (Display)",
    },
    "client": {
        "Start of Day Balance (Local)": "Start of Day Balance (Display)",
        "Credits (Local)": "Credits (Display)",
        "Debits (Local)": "Debits (Display)",
        "Calculated Balance (Local)": "Calculated Balance (Display)",
        "Swing (Local)": "Swing (Display)",
        # No display twin exists for these two in the source data.
        "EOD Balance (Local)": None,
        "Difference (Local)": None,
    },
    "business_ledger": {
        "Amount (Local)": "Amount (Display)",
    },
}

# Aggregations that add values together. A minimum or a maximum picks an
# existing row rather than combining rows, so it stays valid across currencies.
ADDITIVE_AGGREGATIONS = {"sum", "mean", "median", "std", "var",
                         "p25", "p50", "p75", "p90", "p95", "p99"}

# The timestamp that orders events within a day, per view.
TIMESTAMP_COLUMN = {
    "nostro_transfer": "Created Time",
    "client": "Last Transaction Received",
    "business_ledger": "Transaction Timestamp",
}

# Where direction lives in a separate column, so the amount is unsigned.
SIGN_COLUMN = {
    "business_ledger": {"column": "Debit/Credit Mark", "negative": "DR"},
}

# Granularities that describe activity within a day rather than across days.
INTRADAY_GRANULARITIES = {"minute", "15min", "30min", "hour"}

# A peak table is sorted worst-first, so the answer is at the top; showing every
# day for every currency buries it.
PEAK_DEFAULT_LIMIT = 20

# A management breakdown a person reads is a few rows. Past this the answer is
# buried in its own detail, so the tail is folded into a single "Other" row.
MANAGEMENT_ROW_CAP = 25

# Aggregations where folding a tail into "Other" is arithmetically sound. An
# average of averages is not an average, so those are capped rather than folded.
FOLDABLE_AGGREGATIONS = {"sum", "count"}


# Column carrying the primary monetary measure for each view, used when the
# caller asks to aggregate but does not name a measure.
DEFAULT_MEASURE = {
    "nostro_transfer": "Value Amount",
    "client": "Calculated Balance (Display)",
    "business_ledger": "Amount (Display)",
}


class QueryError(ValueError):
    """Raised when a requested query cannot be satisfied by the data."""


def _clean_frame(rows: list[list[Any]], header: list[Any]) -> pd.DataFrame:
    cols = [str(c).strip() for c in header]
    df = pd.DataFrame(rows, columns=cols)
    # Drop the trailing "Total" footer row and any fully-empty rows.
    if len(df) and str(df.iloc[-1, 0]).strip() == TOTAL_ROW_MARKER:
        df = df.iloc[:-1]
    df = df.dropna(how="all")
    return df.reset_index(drop=True)


@lru_cache(maxsize=1)
def _workbook_frames() -> dict[str, pd.DataFrame]:
    """Load the three view sheets. data_only=True reads Excel's cached formula
    results, so derived columns (balances, FX translation, Match) are available
    without needing Excel installed."""
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    frames: dict[str, pd.DataFrame] = {}
    try:
        for view, sheet in VIEW_SHEETS.items():
            ws = wb[sheet]
            all_rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
            frames[view] = _clean_frame([list(r) for r in all_rows[1:]], list(all_rows[0]))
    finally:
        wb.close()
    return frames


REFERENCE_FILE = Path(os.getenv("REFERENCE_FILE") or DATA_DIR / "reference_data.xlsx")
REFERENCE_HEADER_ROW = 4  # title, subtitle, blank, then the header


@lru_cache(maxsize=1)
def _reference_frames() -> dict[str, pd.DataFrame]:
    """Reference tables from the data lake, keyed to values in the live data.

    Loaded separately from the transaction views: they answer "what else do we
    know about this thing", not "what happened", and keeping them apart stops a
    lookup table being mistaken for a source of transactions.
    """
    if not REFERENCE_FILE.exists():
        return {}
    wb = openpyxl.load_workbook(REFERENCE_FILE, data_only=True)
    frames: dict[str, pd.DataFrame] = {}
    try:
        for name in wb.sheetnames:
            rows = list(wb[name].iter_rows(min_row=REFERENCE_HEADER_ROW, values_only=True))
            if len(rows) < 2:
                continue
            header = [str(c).strip() for c in rows[0] if c is not None]
            body = [list(r)[:len(header)] for r in rows[1:] if r and r[0] is not None]
            frames[name] = pd.DataFrame(body, columns=header)
    finally:
        wb.close()
    return frames


def reference_tables() -> dict[str, pd.DataFrame]:
    return _reference_frames()


def get_reference(name: str) -> pd.DataFrame:
    frames = _reference_frames()
    if name not in frames:
        raise QueryError(
            f"Unknown reference table '{name}'. Available: "
            f"{', '.join(sorted(frames)) or 'none loaded'}."
        )
    return frames[name].copy()


def reference_summary() -> dict[str, Any]:
    """Compact description for the prompt: the key and what each table adds."""
    out: dict[str, Any] = {}
    for name, df in _reference_frames().items():
        if df.empty:
            continue
        out[name] = {
            "key_column": df.columns[0],
            "rows": len(df),
            "attributes": [c for c in df.columns[1:]],
        }
    return out


@lru_cache(maxsize=1)
def _metadata() -> dict[str, Any]:
    """Read the workbook's own Data Dictionary, View Controls and Reference Data.

    This is what grounds the assistant's limitations narrative: business rules,
    domains and derivation formulas come from the client's own documentation
    rather than being invented by the model.
    """
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    try:
        dd_rows = list(wb["Data Dictionary"].iter_rows(min_row=METADATA_HEADER_ROW, values_only=True))
        dd_header = [str(c).strip() for c in dd_rows[0]]
        data_dictionary = [
            dict(zip(dd_header, r)) for r in dd_rows[1:] if r and r[0] is not None
        ]

        vc_rows = list(wb["View Controls"].iter_rows(min_row=METADATA_HEADER_ROW, values_only=True))
        vc_header = [str(c).strip() for c in vc_rows[0]]
        view_controls = [
            dict(zip(vc_header, r)) for r in vc_rows[1:] if r and r[0] is not None
        ]

        fx: dict[str, float] = {}
        for row in wb["Reference Data"].iter_rows(min_row=3, values_only=True):
            if row and row[0] and isinstance(row[1], (int, float)):
                fx[str(row[0]).strip()] = float(row[1])
    finally:
        wb.close()

    return {
        "data_dictionary": data_dictionary,
        "view_controls": view_controls,
        "fx_rates": fx,
        "data_classification": "Synthetic / non-production",
    }


def get_metadata() -> dict[str, Any]:
    return _metadata()


def get_frame(view: str) -> pd.DataFrame:
    if view not in VIEW_SHEETS:
        raise QueryError(
            f"Unknown view '{view}'. Available views: {', '.join(VIEW_SHEETS)}."
        )
    return _workbook_frames()[view].copy()


def list_columns(view: str) -> list[str]:
    return list(get_frame(view).columns)


def schema_summary() -> dict[str, Any]:
    """Compact schema for the model's system prompt: columns, dtypes, and the
    distinct values of low-cardinality (enum-like) columns."""
    summary: dict[str, Any] = {}
    for view in VIEW_SHEETS:
        df = get_frame(view)
        cols = []
        for col in df.columns:
            series = df[col].dropna()
            entry: dict[str, Any] = {"name": col, "dtype": _friendly_dtype(series)}
            # Surface enum domains so the model filters using real values.
            # Checked via the friendly dtype rather than `== object`, because
            # pandas 3 gives text columns a dedicated StringDtype.
            if entry["dtype"] == "text" and 0 < series.nunique() <= 12:
                members = {str(v) for v in series.unique()}
                entry["values"] = sorted(members)
                # Glyph columns get word aliases, since a model cannot reliably
                # emit characters like the reconciliation tick and cross.
                if members.issubset(set(GLYPH_ALIASES)):
                    entry["filter_using_these_words_instead"] = {
                        GLYPH_ALIASES[g][0]: g for g in sorted(members)
                    }
            cols.append(entry)
        summary[view] = {
            "sheet": VIEW_SHEETS[view],
            "row_count": len(df),
            "default_measure": DEFAULT_MEASURE[view],
            "columns": cols,
        }
    return summary


def _friendly_dtype(series: pd.Series) -> str:
    """Classify by pandas dtype, not by isinstance on a sample: numpy scalars do
    not subclass the matching Python types (np.int64 is not an int)."""
    if series.empty:
        return "unknown (all values empty)"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    if pd.api.types.is_numeric_dtype(series) and not pd.api.types.is_bool_dtype(series):
        return "number"
    sample = series.iloc[0]
    if isinstance(sample, dt.datetime):
        return "datetime"
    if isinstance(sample, dt.date):
        return "date"
    if isinstance(sample, (int, float)) and not isinstance(sample, bool):
        return "number"
    return "text"


def _coerce_comparable(value: Any, series: pd.Series) -> Any:
    """Align a filter value with the column's runtime type so comparisons work."""
    if series.empty:
        return value
    sample = series.dropna()
    if sample.empty:
        return value
    sample = sample.iloc[0]
    if isinstance(sample, (dt.datetime, dt.date)) and isinstance(value, str):
        try:
            return pd.to_datetime(value).to_pydatetime()
        except (ValueError, TypeError) as exc:
            raise QueryError(f"Could not read '{value}' as a date.") from exc
    if isinstance(sample, (int, float)) and isinstance(value, str):
        try:
            return float(value)
        except ValueError as exc:
            raise QueryError(f"Could not read '{value}' as a number.") from exc
    return value


def _normalise_glyph_value(value: Any, series: pd.Series) -> Any:
    """Map a word like 'unmatched' onto the glyph actually stored in the column.

    Only applies to columns whose values are glyphs, so ordinary text filtering
    is untouched. Also catches the control characters a model sometimes emits
    when it cannot reproduce a glyph: if the column is a glyph column and the
    value matches none of its members, that filter would silently return zero
    rows, which reads as a real finding of 'nothing'.
    """
    members = {str(v) for v in series.dropna().unique()}
    if not members or not members.issubset(set(GLYPH_ALIASES)):
        return value

    text = str(value).strip()
    if text in members:
        return text
    mapped = _ALIAS_LOOKUP.get(text.casefold())
    if mapped in members:
        return mapped
    raise QueryError(
        f"Value {text!r} is not valid for this column. Use one of: "
        + ", ".join(sorted(f"'{m}'" for m in members))
        + " - or the words "
        + ", ".join(sorted({a for g in members for a in GLYPH_ALIASES[g][:2]}))
        + "."
    )


def _filter_mask(df: pd.DataFrame, spec: dict[str, Any]) -> pd.Series:
    """Build a boolean mask for one filter, or for a nested any/all group.

    Masks rather than successive slicing, because "failed OR rejected" cannot be
    expressed by filtering twice - and answering the intersection instead is a
    silent, badly wrong answer rather than an error.
    """
    if "any" in spec or "all" in spec:
        key = "any" if "any" in spec else "all"
        parts = spec.get(key) or []
        if not parts:
            return pd.Series(True, index=df.index)
        masks = [_filter_mask(df, p) for p in parts]
        combined = masks[0]
        for mask in masks[1:]:
            combined = (combined | mask) if key == "any" else (combined & mask)
        return ~combined if spec.get("negate") else combined

    mask = _single_mask(df, spec)
    return ~mask if spec.get("negate") else mask


def _single_mask(df: pd.DataFrame, spec: dict[str, Any]) -> pd.Series:
    col, op = spec.get("column"), (spec.get("operator") or "eq").lower()
    value = spec.get("value")
    if col not in df.columns:
        raise QueryError(
            f"Column '{col}' does not exist. Available columns: {', '.join(df.columns)}."
        )

    series = df[col]
    if op in {"eq", "ne"}:
        value = _normalise_glyph_value(value, series)
    if op in {"in", "not_in"}:
        values = value if isinstance(value, list) else [value]
        values = {str(v).strip().casefold() for v in values}
        mask = series.astype(str).str.strip().str.casefold().isin(values)
        return ~mask if op == "not_in" else mask

    if op == "contains":
        return series.astype(str).str.contains(str(value), case=False, na=False)

    if op == "is_null":
        return series.isna()
    if op == "not_null":
        return series.notna()

    target = _coerce_comparable(value, series)
    ops = {
        "eq": lambda s: s == target,
        "ne": lambda s: s != target,
        "gt": lambda s: s > target,
        "gte": lambda s: s >= target,
        "lt": lambda s: s < target,
        "lte": lambda s: s <= target,
    }
    if op not in ops:
        raise QueryError(f"Unsupported operator '{op}'.")
    # Text columns compared with eq/ne should ignore case and padding.
    if op in {"eq", "ne"} and isinstance(target, str):
        mask = series.astype(str).str.strip().str.casefold() == target.strip().casefold()
        return ~mask if op == "ne" else mask
    return ops[op](series)


AGGREGATIONS = {"sum", "mean", "count", "min", "max", "median", "std", "var",
                "nunique", "p25", "p50", "p75", "p90", "p95", "p99"}

# Aggregations that work on any column, not only a numeric one.
NON_NUMERIC_AGGREGATIONS = {"count", "nunique", "min", "max"}


def _aggregate_series(series: pd.Series, how: str):
    """Apply an aggregation name, including the percentile forms."""
    if how == "nunique":
        return series.nunique(dropna=True)
    if how.startswith("p") and how[1:].isdigit():
        return series.quantile(int(how[1:]) / 100)
    if how == "std":
        return series.std(ddof=1)
    if how == "var":
        return series.var(ddof=1)
    return getattr(series, how)()


def describe_series(series: pd.Series) -> dict[str, float]:
    """The standard distribution summary, computed deterministically."""
    clean = pd.to_numeric(series, errors="coerce").dropna()
    if clean.empty:
        return {}
    mean = float(clean.mean())
    std = float(clean.std(ddof=1)) if len(clean) > 1 else 0.0
    return {
        "count": int(clean.count()),
        "mean": mean,
        "std": std,
        "min": float(clean.min()),
        "p25": float(clean.quantile(0.25)),
        "median": float(clean.median()),
        "p75": float(clean.quantile(0.75)),
        "p95": float(clean.quantile(0.95)),
        "max": float(clean.max()),
        # Reported so a reader can see how far the tails actually reach, rather
        # than assuming a normal distribution.
        "mean_minus_3std": mean - 3 * std,
        "mean_plus_3std": mean + 3 * std,
        "skew": float(clean.skew()) if len(clean) > 2 else 0.0,
        "within_1std_pct": float(((clean - mean).abs() <= std).mean() * 100) if std else 100.0,
        "within_2std_pct": float(((clean - mean).abs() <= 2 * std).mean() * 100) if std else 100.0,
        "within_3std_pct": float(((clean - mean).abs() <= 3 * std).mean() * 100) if std else 100.0,
    }

# Pandas offset aliases for time bucketing, keyed by the words a user would use.
TIME_GRANULARITIES = {
    "minute": ("min", "%Y-%m-%d %H:%M"),
    "15min": ("15min", "%Y-%m-%d %H:%M"),
    "30min": ("30min", "%Y-%m-%d %H:%M"),
    "hour": ("h", "%Y-%m-%d %H:00"),
    "day": ("D", "%Y-%m-%d"),
    "week": ("W", "%Y-%m-%d"),
    "month": ("MS", "%Y-%m"),
}


def apply_join(df: pd.DataFrame, spec: dict[str, Any], base_view: str) -> tuple[pd.DataFrame, dict]:
    """Look up columns from another view - vlookup semantics.

    A join that matches nothing must say so. Returning an empty frame would read
    as a genuine finding of "no records", when the truth is that the two key
    columns hold different populations entirely.
    """
    right_view = spec.get("view")
    references = _reference_frames()
    if right_view in references:
        right = get_reference(right_view)
        right_label = right_view
    elif right_view in VIEW_SHEETS:
        right = get_frame(right_view)
        right_label = VIEW_SHEETS[right_view]
    else:
        raise QueryError(
            f"Cannot combine with '{right_view}'. Available views: "
            f"{', '.join(VIEW_SHEETS)}. Available reference tables: "
            f"{', '.join(sorted(references)) or 'none'}."
        )

    on = spec.get("on") or {}
    left_key = on.get("left")
    # A reference table has one key, so the caller need not name it.
    right_key = on.get("right") or (right.columns[0] if right_view in references
                                    else on.get("left"))
    if not left_key or not right_key:
        raise QueryError("A combine needs a key column on each side.")

    _require_column(df, left_key, "a join key", base_view)
    if right_key not in right.columns:
        raise QueryError(
            f"Column '{right_key}' does not exist in {right_label}. "
            f"Available: {', '.join(right.columns)}."
        )

    bring = [c for c in (spec.get("bring") or []) if c]
    for col in bring:
        if col not in right.columns:
            raise QueryError(
                f"Cannot bring '{col}' from {right_label}. "
                f"Available: {', '.join(right.columns)}."
            )
    if not bring:
        bring = [c for c in right.columns if c != right_key][:4]

    # The key is often listed among the columns to bring, which is a reasonable
    # thing to ask for and would otherwise select it twice and break the merge.
    bring = [c for c in dict.fromkeys(bring) if c != right_key]
    if not bring:
        # Asking for only the key is a harmless mistake - the base data already
        # has it. Bring the whole record rather than refusing over wording.
        bring = [c for c in right.columns if c != right_key][:6]

    left_keys = df[left_key].astype(str).str.strip()
    # vlookup takes the first match, so collapse duplicates rather than
    # multiplying rows.
    right_slim = right[[right_key] + bring].copy()
    right_slim[right_key] = right_slim[right_key].astype(str).str.strip()
    right_slim = right_slim.drop_duplicates(subset=[right_key], keep="first")

    overlap = set(left_keys) & set(right_slim[right_key])
    if not overlap:
        raise QueryError(
            f"'{left_key}' in {VIEW_SHEETS[base_view]} and '{right_key}' in "
            f"{right_label} have no values in common, so they cannot "
            f"be combined. Examples: {sorted(set(left_keys))[:2]} against "
            f"{sorted(set(right_slim[right_key]))[:2]}. These views do not share "
            "a key in this dataset."
        )

    # Prefix collisions so an existing column is never silently overwritten.
    renames = {c: (f"{c} ({right_view})" if c in df.columns else c) for c in bring}
    right_slim = right_slim.rename(columns=renames)

    merged = df.assign(__key__=left_keys).merge(
        right_slim.rename(columns={right_key: "__key__"}), on="__key__", how="left"
    ).drop(columns="__key__")

    brought = list(renames.values())
    matched = int(merged[brought[0]].notna().sum()) if brought else 0
    return merged, {
        "view": right_view,
        "on": f"{left_key} = {right_key}",
        "brought": brought,
        "rows_matched": matched,
        "rows_unmatched": len(merged) - matched,
    }


AGE_UNITS = {"minutes": 60.0, "hours": 3600.0, "days": 86400.0}
DEFAULT_AGE_BANDS = [1, 4, 24]  # hours


def data_as_of(view: str) -> dt.datetime | None:
    """The latest timestamp in a view - the point the data is current to.

    Ageing is measured against this rather than the wall clock: the data is an
    extract, and against today's date every historical row would look equally
    stale, which tells a reader nothing about the queue when it was captured.
    """
    df = get_frame(view)
    latest = None
    for col in df.columns:
        if not ("time" in col.lower() or "timestamp" in col.lower()):
            continue
        stamps = pd.to_datetime(df[col], errors="coerce").dropna()
        if not stamps.empty:
            top = stamps.max()
            latest = top if latest is None else max(latest, top)
    return latest.to_pydatetime() if latest is not None else None


def apply_age(df: pd.DataFrame, spec: dict[str, Any], view: str,
              reference: dt.datetime | None) -> tuple[pd.DataFrame, list[str]]:
    """Add elapsed time since a timestamp, optionally banded.

    "How long has this been waiting" is the actionable part of any queue view;
    who created it is not.
    """
    column = spec.get("age_of")
    _require_column(df, column, "an ageing column", view)
    stamps = pd.to_datetime(df[column], errors="coerce")
    if stamps.notna().sum() == 0:
        raise QueryError(f"Column '{column}' does not contain readable timestamps.")

    if reference is None:
        reference = stamps.max().to_pydatetime()

    unit = str(spec.get("unit") or "hours").lower()
    if unit not in AGE_UNITS:
        raise QueryError(f"Unsupported age unit '{unit}'. Use one of: {', '.join(AGE_UNITS)}.")

    elapsed = (reference - stamps).dt.total_seconds() / AGE_UNITS[unit]
    name = spec.get("name") or f"Age ({unit})"
    added = [name]
    df = df.assign(**{name: elapsed.round(2)})

    bands = spec.get("bands")
    if bands is not False:
        edges = sorted(float(b) for b in (bands or DEFAULT_AGE_BANDS))
        # Band on hours regardless of the numeric unit, since that is how an
        # operations team talks about a queue.
        hours = (reference - stamps).dt.total_seconds() / 3600.0
        labels, cuts = [], [-float("inf")] + edges + [float("inf")]
        for i in range(len(cuts) - 1):
            low, high = cuts[i], cuts[i + 1]
            if low == -float("inf"):
                labels.append(f"under {high:g}h")
            elif high == float("inf"):
                labels.append(f"over {low:g}h")
            else:
                labels.append(f"{low:g}-{high:g}h")
        band_name = f"{name} band"
        df = df.assign(**{band_name: pd.cut(hours, bins=cuts, labels=labels,
                                            right=False).astype(str)})
        added.append(band_name)
    return df, added


TIME_PARTS = {
    "hour_of_day": ("Hour of day", lambda s: s.dt.hour.map(lambda h: f"{h:02d}:00")),
    "weekday": ("Day of week", lambda s: s.dt.day_name()),
    "day_of_month": ("Day of month", lambda s: s.dt.day),
    "week_of_year": ("Week", lambda s: s.dt.isocalendar().week.astype(int)),
    "month": ("Month", lambda s: s.dt.strftime("%Y-%m")),
}
WEEKDAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                 "Saturday", "Sunday"]


def apply_time_part(df: pd.DataFrame, spec: dict[str, Any], view: str) -> pd.DataFrame:
    """Pull a repeating part out of a timestamp - the hour of the day, the day
    of the week - so it can be grouped or filtered on.

    A time bucket answers "when did this happen"; a time part answers "does this
    happen at the same point in every day or week", which is a different
    question and cannot be reached by bucketing.
    """
    column = spec.get("part_of")
    _require_column(df, column, "a time part", view)
    part = str(spec.get("part") or "hour_of_day").lower()
    if part not in TIME_PARTS:
        raise QueryError(f"Unsupported time part '{part}'. "
                         f"Use one of: {', '.join(TIME_PARTS)}.")

    stamps = pd.to_datetime(df[column], errors="coerce")
    if stamps.notna().sum() == 0:
        raise QueryError(f"Column '{column}' does not contain readable timestamps.")

    default_name, extract = TIME_PARTS[part]
    name = spec.get("name") or default_name
    return df.assign(**{name: extract(stamps)})


def profile_completeness(view: str, df: pd.DataFrame,
                         columns: list[str] | None) -> pd.DataFrame:
    """How populated each column is.

    Every other view inherits the quality of these fields, so this is the one
    report whose entire value is in what is absent.
    """
    total = len(df)
    wanted = [c for c in (columns or df.columns) if c in df.columns]
    rows = []
    for col in wanted:
        series = df[col]
        blank = series.isna()
        if not pd.api.types.is_numeric_dtype(series):
            blank = blank | series.astype(str).str.strip().isin(["", "None", "nan", "NaT"])
        missing = int(blank.sum())
        rows.append({
            "Column": col,
            "Populated": total - missing,
            "Missing": missing,
            "% populated": round(100 * (total - missing) / total, 1) if total else 0.0,
            "Distinct values": int(series.nunique(dropna=True)),
        })
    result = pd.DataFrame(rows).sort_values("% populated").reset_index(drop=True)
    return result


def apply_duration(df: pd.DataFrame, spec: dict[str, Any], view: str) -> pd.DataFrame:
    """Elapsed time between two timestamp columns, in a stated unit.

    Subtracting two timestamps with ordinary arithmetic yields nanoseconds -
    a turnaround of "2,040,000,000" that is really 34 minutes, and no label to
    say so. The unit is always part of the column name here.
    """
    start, end = spec.get("duration_from"), spec.get("duration_to")
    _require_column(df, start, "the start of a duration", view)
    _require_column(df, end, "the end of a duration", view)

    unit = str(spec.get("unit") or "hours").lower()
    if unit not in AGE_UNITS:
        raise QueryError(f"Unsupported duration unit '{unit}'. "
                         f"Use one of: {', '.join(AGE_UNITS)}.")

    begin = pd.to_datetime(df[start], errors="coerce")
    finish = pd.to_datetime(df[end], errors="coerce")
    if begin.notna().sum() == 0 or finish.notna().sum() == 0:
        raise QueryError(
            f"A duration needs two timestamp columns; '{start}' and '{end}' do "
            "not both contain readable timestamps."
        )

    name = spec.get("name") or f"Duration ({unit})"
    if unit not in name.lower():
        name = f"{name} ({unit})"
    elapsed = (finish - begin).dt.total_seconds() / AGE_UNITS[unit]
    return df.assign(**{name: elapsed.round(2)})


def _is_temporal(series: pd.Series) -> bool:
    if pd.api.types.is_datetime64_any_dtype(series):
        return True
    sample = series.dropna()
    return bool(len(sample)) and isinstance(sample.iloc[0], (dt.datetime, dt.date))


def apply_derived(df: pd.DataFrame, specs: list[dict[str, Any]], view: str) -> tuple[pd.DataFrame, list[str]]:
    """Add calculated columns from arithmetic between two columns, or a column
    and a constant."""
    added: list[str] = []
    ops = {
        "+": lambda a, b: a + b, "-": lambda a, b: a - b,
        "*": lambda a, b: a * b, "/": lambda a, b: a / b.replace(0, pd.NA),
    }
    reference = data_as_of(view)
    for spec in specs or []:
        name = spec.get("name") or "Derived"
        if spec.get("age_of"):
            df, age_columns = apply_age(df, spec, view, reference)
            added.extend(age_columns)
            continue
        if spec.get("duration_from") and spec.get("duration_to"):
            before = set(df.columns)
            df = apply_duration(df, spec, view)
            added.extend(sorted(set(df.columns) - before))
            continue
        if spec.get("part_of"):
            before = set(df.columns)
            df = apply_time_part(df, spec, view)
            added.extend(sorted(set(df.columns) - before))
            continue
        op = spec.get("op")
        if op not in ops:
            raise QueryError(
                f"'{op}' is not a calculation. A derived column combines two "
                f"columns arithmetically ({', '.join(ops)}). To select rows by a "
                "comparison use a filter instead, or count distinct values with "
                "the nunique aggregation."
            )
        left, right = spec.get("left"), spec.get("right")
        _require_column(df, left, "a calculation input", view)

        # Timestamps must never be coerced into numbers. Doing so turns a
        # subtraction into nanoseconds, which reads as a plausible figure and is
        # wrong by nine orders of magnitude.
        if _is_temporal(df[left]) or (isinstance(right, str) and right in df.columns
                                      and _is_temporal(df[right])):
            if op == "-" and isinstance(right, str) and right in df.columns:
                raise QueryError(
                    f"'{left}' and '{right}' are timestamps. To measure the time "
                    "between them use a duration - give duration_from, "
                    "duration_to and a unit - rather than subtracting them, "
                    "which would produce nanoseconds."
                )
            raise QueryError(
                f"'{left}' is a timestamp and cannot be used in arithmetic. "
                "Use a duration between two timestamps, or an age from one."
            )

        df = _as_numeric(df, left)
        left_series = df[left]

        if isinstance(right, (int, float)):
            right_series = pd.Series([float(right)] * len(df), index=df.index)
        else:
            _require_column(df, right, "a calculation input", view)
            df = _as_numeric(df, right)
            right_series = df[right]

        df = df.assign(**{name: ops[op](left_series, right_series)})
        added.append(name)
    return df, added


def _cap_for_reading(result: pd.DataFrame, label_columns: list[str],
                     measure_columns: list[str], measures: list[dict[str, str]],
                     is_time_series: bool) -> tuple[pd.DataFrame, str | None]:
    """Keep a management breakdown to a readable length.

    A time series is left alone: its rows are periods, and folding the tail
    would mean folding "later", which is meaningless. Everything else is sorted
    by size already, so the tail is genuinely the small stuff - summed into a
    single "Other" row where that is sound, capped where it is not.
    """
    if is_time_series or len(result) <= MANAGEMENT_ROW_CAP or not label_columns:
        return result, None

    total_groups = len(result)
    keep = MANAGEMENT_ROW_CAP - 1
    head, tail = result.head(keep).copy(), result.iloc[keep:]

    aggs = {m.get("aggregation", "sum") for m in measures}
    if aggs <= FOLDABLE_AGGREGATIONS:
        other = {label_columns[0]: f"Other ({len(tail)})"}
        for col in label_columns[1:]:
            other[col] = ""
        for col in measure_columns:
            other[col] = tail[col].sum() if col in tail.columns else None
        folded = pd.concat([head, pd.DataFrame([other])], ignore_index=True)
        return folded, (
            f"{total_groups} groups were returned. The {keep} largest are shown "
            f"individually and the remaining {len(tail)} are combined into "
            '"Other" - a breakdown this long is not readable as a management '
            "view. Ask for a specific group, or a top-N, to go further."
        )

    return head, (
        f"Showing the {keep} largest of {total_groups} groups. The rest are "
        f"omitted rather than combined, because averaging an average would not "
        "give a meaningful figure."
    )


def _has_date_filter(filters: list[dict[str, Any]], df: pd.DataFrame) -> bool:
    """Whether the caller already narrowed the period themselves."""
    def touches_date(spec: dict[str, Any]) -> bool:
        if "any" in spec or "all" in spec:
            return any(touches_date(p) for p in spec.get("any") or spec.get("all") or [])
        col = str(spec.get("column") or "")
        if col not in df.columns:
            return False
        return ("date" in col.lower() or "time" in col.lower()
                or pd.api.types.is_datetime64_any_dtype(df[col]))
    return any(touches_date(f) for f in filters)


def compute_peak(df: pd.DataFrame, view: str, measure: str, timestamp_col: str,
                 group_by: list[str], signed_by: dict[str, Any] | None = None
                 ) -> pd.DataFrame:
    """Peak intraday position: the extremes of the running net position, per day.

    This is the shape of the BCBS 248 monitoring metric. A running total that
    never resets describes a month-long accumulation, not an intraday position -
    the position starts each day from the opening balance, so the cumulative
    must restart every day and the peak is the extreme reached within it.

    Returns one row per group per day: the largest positive position, the
    largest negative (which is the usage figure), when each occurred, and where
    the day closed.
    """
    _require_column(df, measure, "the peak measure", view)
    _require_column(df, timestamp_col, "the peak timestamp", view)
    if df.empty:
        # Blaming the timestamp column here sends the reader to investigate the
        # wrong thing: the column is fine, the filters simply matched nothing.
        raise QueryError(
            "No records matched the filters, so there is no intraday position to "
            "measure. Widen the filters, or check that the values filtered on "
            "actually occur in this data."
        )
    df = _as_numeric(df, measure)

    stamps = pd.to_datetime(df[timestamp_col], errors="coerce")
    if stamps.notna().sum() == 0:
        raise QueryError(f"Column '{timestamp_col}' does not contain readable timestamps.")

    work = df.assign(__stamp__=stamps, __day__=stamps.dt.date)
    amount = work[measure]

    # A debit/credit column means the amount carries no sign of its own.
    if signed_by:
        col, negative = signed_by.get("column"), str(signed_by.get("negative", "DR"))
        if col in work.columns:
            sign = work[col].astype(str).str.strip().str.upper().eq(negative.upper())
            amount = amount.abs() * sign.map({True: -1, False: 1})
    work = work.assign(__amt__=amount)

    partition = ["__day__"] + [g for g in group_by if g in work.columns]
    work = work.sort_values(partition + ["__stamp__"]).reset_index(drop=True)
    work["__cum__"] = work.groupby(partition, sort=False)["__amt__"].cumsum()

    rows = []
    for key, part in work.groupby(partition, sort=False):
        key = key if isinstance(key, tuple) else (key,)
        peak_idx, trough_idx = part["__cum__"].idxmax(), part["__cum__"].idxmin()
        record = {"Value Date": key[0]}
        for name, value in zip(partition[1:], key[1:]):
            record[name] = value
        record.update({
            "Peak position": round(float(part.loc[peak_idx, "__cum__"]), 3),
            "Peak at": part.loc[peak_idx, "__stamp__"].strftime("%H:%M"),
            "Largest usage": round(float(part.loc[trough_idx, "__cum__"]), 3),
            "Usage at": part.loc[trough_idx, "__stamp__"].strftime("%H:%M"),
            "Closing position": round(float(part["__cum__"].iloc[-1]), 3),
            "Movements": int(len(part)),
        })
        rows.append(record)

    result = pd.DataFrame(rows)
    # Worst usage first: the largest negative position is the number that
    # matters for an intraday liquidity requirement.
    return result.sort_values("Largest usage").reset_index(drop=True)


def compute_backlog(df: pd.DataFrame, view: str, opened_col: str, closed_col: str,
                    granularity: str, measure: str | None = None) -> pd.DataFrame:
    """How a queue evolved: what was outstanding at each point in time.

    A queue banded by age answers "how bad is it now". It cannot answer "how did
    it get like this", which is the question a manager actually acts on - a
    backlog that built steadily needs more people, one that spiked at 14:00
    needs to know what happened at 14:00. Both look identical as a single bar.

    An item is outstanding at time t if it arrived on or before t and had not
    yet cleared. Items that never cleared are outstanding from arrival onwards.
    """
    _require_column(df, opened_col, "the queue arrival time", view)
    _require_column(df, closed_col, "the queue clearance time", view)
    if granularity not in TIME_GRANULARITIES:
        raise QueryError(f"Unsupported interval '{granularity}'. "
                         f"Use one of: {', '.join(TIME_GRANULARITIES)}.")

    opened = pd.to_datetime(df[opened_col], errors="coerce")
    closed = pd.to_datetime(df[closed_col], errors="coerce")
    if opened.notna().sum() == 0:
        raise QueryError(f"'{opened_col}' does not contain readable timestamps.")

    valid = opened.notna()
    opened, closed = opened[valid], closed[valid]
    amounts = None
    if measure and measure in df.columns:
        amounts = pd.to_numeric(df.loc[valid, measure], errors="coerce").fillna(0).abs()

    freq, fmt = TIME_GRANULARITIES[granularity]
    start = opened.min().floor(freq if freq not in {"W", "MS"} else "D")
    last_event = max(opened.max(), closed.max()) if closed.notna().any() else opened.max()
    grid = pd.date_range(start, last_event.ceil(freq if freq not in {"W", "MS"} else "D"),
                         freq=freq)
    if len(grid) > 400:
        raise QueryError(
            f"That period at {granularity} intervals would produce {len(grid)} "
            "points, which is not readable. Narrow to a single day, or use a "
            "coarser interval such as hour."
        )

    rows = []
    for point in grid:
        arrived = opened <= point
        cleared = closed.notna() & (closed <= point)
        outstanding = arrived & ~cleared
        row = {
            f"{opened_col} ({granularity})": point.strftime(fmt),
            "Arrived": int((opened.dt.floor(freq) == point).sum()),
            "Cleared": int((closed.dt.floor(freq) == point).sum()) if closed.notna().any() else 0,
            "Outstanding": int(outstanding.sum()),
        }
        if amounts is not None:
            row["Outstanding value"] = round(float(amounts[outstanding].sum()), 3)
        rows.append(row)

    result = pd.DataFrame(rows)
    # Leading and trailing stretches with nothing outstanding are not the queue,
    # they are the hours around it.
    active = result["Outstanding"] > 0
    if active.any():
        result = result.loc[active.idxmax():active[::-1].idxmax()].reset_index(drop=True)
    return result


def guard_currency(view: str, df: pd.DataFrame, measures: list[dict[str, str]],
                   group_by: list[str]) -> tuple[list[dict[str, str]], list[str]]:
    """Stop local-currency amounts being added together across currencies.

    Adding 1 JPY to 1 GBP produces a number that means nothing, and a caveat
    saying so does not prevent a reader acting on it - so the query is corrected
    to the FX-translated column, or refused when no such column exists.

    Grouping by the currency column makes each group single-currency, so local
    amounts stay valid and nothing is changed.
    """
    ccy_col = CURRENCY_COLUMN.get(view)
    if not ccy_col or ccy_col not in df.columns:
        return measures, []
    # Grouping by currency means every group holds one currency.
    if any(g == ccy_col for g in group_by or []):
        return measures, []
    if df[ccy_col].nunique(dropna=True) <= 1:
        return measures, []

    equivalents = DISPLAY_EQUIVALENT.get(view, {})
    corrected: list[dict[str, str]] = []
    notes: list[str] = []
    spanning = df[ccy_col].nunique(dropna=True)

    for measure in measures:
        column = measure.get("column")
        agg = (measure.get("aggregation") or "sum").lower()
        if column not in equivalents or agg not in ADDITIVE_AGGREGATIONS:
            corrected.append(measure)
            continue

        display = equivalents[column]
        if display and display in df.columns:
            swapped = dict(measure)
            swapped["column"] = display
            swapped["label"] = display
            corrected.append(swapped)
            notes.append(
                f"'{column}' is held in each row's own currency and this result "
                f"spans {spanning} currencies, so adding it would combine unlike "
                f"units. '{display}' was used instead, which is FX-translated to a "
                "single currency."
            )
        else:
            raise QueryError(
                f"'{column}' is denominated in each row's own currency and this "
                f"query spans {spanning} currencies, so the values cannot be "
                f"combined - the total would add unlike units. Either group by "
                f"'{ccy_col}' so each figure covers one currency, filter to a "
                f"single currency, or use an FX-translated column. "
                f"'{column}' has no FX-translated equivalent in this data."
            )
    return corrected, notes


def _require_column(df: pd.DataFrame, col: str, role: str, view: str) -> None:
    if col not in df.columns:
        raise QueryError(
            f"Cannot use '{col}' as {role} - no such column. "
            f"Available columns: {', '.join(get_frame(view).columns)}."
        )


def _as_numeric(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if pd.api.types.is_numeric_dtype(df[col]):
        return df
    coerced = pd.to_numeric(df[col], errors="coerce")
    if coerced.notna().sum() == 0:
        raise QueryError(f"Column '{col}' is not numeric, so it cannot be aggregated.")
    return df.assign(**{col: coerced})


def _apply_time_bucket(
    df: pd.DataFrame, spec: dict[str, Any], view: str
) -> tuple[pd.DataFrame, str]:
    """Floor a timestamp column into buckets and return the new label column.

    Without this, grouping by a raw timestamp yields one group per row, which
    looks like a time series but carries no information.
    """
    col = spec.get("column")
    _require_column(df, col, "a time bucket", view)
    gran = str(spec.get("granularity") or "hour").lower()
    if gran not in TIME_GRANULARITIES:
        raise QueryError(
            f"Unsupported time granularity '{gran}'. "
            f"Use one of: {', '.join(TIME_GRANULARITIES)}."
        )
    freq, fmt = TIME_GRANULARITIES[gran]

    stamps = pd.to_datetime(df[col], errors="coerce")
    if stamps.notna().sum() == 0:
        raise QueryError(f"Column '{col}' does not contain readable timestamps.")

    # A date with no time component cannot be bucketed within a day: every value
    # would land at midnight and collapse into a single row that looks like an
    # answer.
    if gran in INTRADAY_GRANULARITIES:
        clock = stamps.dropna()
        if ((clock.dt.hour == 0) & (clock.dt.minute == 0)
                & (clock.dt.second == 0)).all():
            raise QueryError(
                f"'{col}' holds a date with no time of day, so it cannot be "
                f"broken down by {gran} - every row would fall at midnight. Use "
                "a column that carries a timestamp, or a 'day' granularity."
            )

    label = f"{col} ({gran})"
    bucketed = stamps.dt.floor(freq) if freq not in {"W", "MS"} else stamps.dt.to_period(
        "W" if freq == "W" else "M"
    ).dt.start_time
    return df.assign(**{label: bucketed.dt.strftime(fmt)}), label


def apply_rate(df: pd.DataFrame, result: pd.DataFrame, group_by: list[str],
               spec: dict[str, Any]) -> tuple[pd.DataFrame, str]:
    """Share of rows in each group meeting a condition.

    A count of failures ranks the busiest venues, not the least reliable ones.
    Three failures out of five is a different problem from ten out of a
    thousand, and only a rate distinguishes them.
    """
    condition = spec.get("where")
    if not condition:
        raise QueryError("A rate needs a 'where' condition to measure against.")

    mask = _filter_mask(df, condition)
    totals = df.groupby(group_by, dropna=False).size().rename("__total__")
    hits = df[mask].groupby(group_by, dropna=False).size().rename("__hits__")
    combined = pd.concat([totals, hits], axis=1).fillna(0).reset_index()

    name = spec.get("name") or "Rate %"
    combined[name] = (combined["__hits__"] / combined["__total__"] * 100).round(1)
    combined["Of total"] = combined["__total__"].astype(int)
    combined["Matching"] = combined["__hits__"].astype(int)

    merged = result.merge(
        combined[group_by + [name, "Matching", "Of total"]], on=group_by, how="left")
    return merged, name


def _normalise_measures(
    measures: list[dict[str, Any]] | None, view: str, df: pd.DataFrame
) -> list[dict[str, str]]:
    """Coerce the measure list into [{column, aggregation, label}]."""
    out: list[dict[str, str]] = []
    for m in measures or []:
        if isinstance(m, str):
            m = {"column": m}
        agg = str(m.get("aggregation") or "sum").lower()
        if agg not in AGGREGATIONS:
            raise QueryError(f"Unsupported aggregation '{agg}'.")
        col = m.get("column")
        if agg == "count" and not col:
            out.append({"column": "", "aggregation": "count", "label": "count"})
            continue
        _require_column(df, col, "a measure", view)
        if agg == "count":
            # Counting a named column is a count of its populated values, and
            # must not be labelled as the column itself - that produced a header
            # promising dates above a column of counts, and an empty one beside it.
            out.append({"column": col, "aggregation": "count",
                        "label": f"count of {col}"})
            continue
        out.append({"column": col, "aggregation": agg, "label": col})

    # Two aggregations of the same column would collide on the column name, so
    # qualify those labels with the aggregation ("p95 of Amount", "median of...").
    counts: dict[str, int] = {}
    for m in out:
        counts[m["label"]] = counts.get(m["label"], 0) + 1
    for m in out:
        if counts[m["label"]] > 1 and m.get("column"):
            m["label"] = f"{m['aggregation']} of {m['column']}"

    # Qualifying is not always enough - two bare counts both qualify to the same
    # thing - so guarantee uniqueness rather than letting pandas raise on a
    # duplicate column label.
    seen: dict[str, int] = {}
    for m in out:
        base = m["label"]
        if base in seen:
            seen[base] += 1
            m["label"] = f"{base} ({seen[base]})"
        else:
            seen[base] = 1
    return out


def run_query(
    view: str,
    mode: str = "aggregate",
    filters: list[dict[str, Any]] | None = None,
    group_by: list[str] | None = None,
    time_bucket: dict[str, Any] | None = None,
    measures: list[dict[str, Any]] | None = None,
    columns: list[str] | None = None,
    sort_by: str | None = None,
    sort_desc: bool = True,
    limit: int | None = None,
    join: dict[str, Any] | list[dict[str, Any]] | None = None,
    derived: list[dict[str, Any]] | None = None,
    rate: dict[str, Any] | None = None,
    backlog: dict[str, Any] | None = None,
    add_share_of_total: bool = False,
    add_cumulative: bool = False,
    # Accepted for convenience; folded into `measures`.
    measure: str | None = None,
    aggregation: str | None = None,
) -> dict[str, Any]:
    """Run a deterministic query and return the result plus a provenance record.

    Two modes:
      aggregate - group and summarise (optionally bucketing a timestamp)
      list      - return matching rows as they are, sorted and limited
    """
    df = get_frame(view)
    total_rows = len(df)

    # Order matters: bring in looked-up columns first, then calculate from them,
    # so both are available to filter on.
    # Joins may be chained: the second hop keys off a column the first brought
    # in. That is how a question reaches an attribute two steps away - a
    # counterparty's country, and then that country's sanctions regime.
    join_info: list[dict[str, Any]] | dict[str, Any] | None = None
    if join:
        hops = join if isinstance(join, list) else [join]
        collected = []
        for hop in hops:
            df, info = apply_join(df, hop, view)
            collected.append(info)
        join_info = collected if len(collected) > 1 else collected[0]

    derived_columns: list[str] = []
    if derived:
        df, derived_columns = apply_derived(df, derived, view)

    filters = filters or []
    if filters:
        mask = pd.Series(True, index=df.index)
        for spec in filters:
            mask &= _filter_mask(df, spec)
        df = df[mask]

    # An intraday question means one day. Bucketing a month by hour produces
    # hundreds of points that read as noise rather than a profile, so scope to
    # the most recent day present unless the caller has already chosen a period.
    scope_note: str | None = None
    if (time_bucket
            and str(time_bucket.get("granularity") or "").lower() in INTRADAY_GRANULARITIES
            and not _has_date_filter(filters, df)
            and not df.empty):
        stamp_col = time_bucket.get("column")
        if stamp_col in df.columns:
            stamps = pd.to_datetime(df[stamp_col], errors="coerce")
            days = stamps.dt.date.dropna()
            if days.nunique() > 1:
                latest = days.max()
                df = df[stamps.dt.date == latest]
                scope_note = (
                    f"An hourly profile spanning {days.nunique()} days is not an "
                    f"intraday view, so this covers the most recent day in the "
                    f"data ({latest:%d %b %Y}). Ask for a specific date, or for a "
                    "daily breakdown, to see the whole period."
                )
    rows_after_filters = len(df)

    mode = (mode or "aggregate").lower()
    if mode not in {"aggregate", "list", "distribution", "peak", "quality",
                    "backlog"}:
        raise QueryError(
            f"Unsupported mode '{mode}'. Use 'aggregate', 'list', "
            "'distribution', 'peak', 'quality' or 'backlog'."
        )

    # Fold the single-measure convenience form into the list form.
    if measure or aggregation:
        measures = measures or []
        if not measures:
            measures = [{"column": measure, "aggregation": aggregation or "sum"}]

    label_columns: list[str] = []
    measure_columns: list[str] = []
    measures_note: str | None = None
    currency_notes: list[str] = []

    raw_values: pd.DataFrame | None = None

    if mode == "backlog":
        spec = backlog or {}
        opened_col = spec.get("opened_at") or TIMESTAMP_COLUMN.get(view)
        closed_col = spec.get("closed_at")
        if not closed_col:
            raise QueryError(
                "A backlog needs both the time an item arrived and the time it "
                "cleared. Give 'opened_at' and 'closed_at'."
            )
        resolved_measures = _normalise_measures(measures, view, df)
        measure_col = resolved_measures[0]["column"] if resolved_measures else None
        result = compute_backlog(
            df, view, opened_col, closed_col,
            str(spec.get("granularity") or "15min").lower(), measure_col)
        label_columns = [result.columns[0]]
        measure_columns = [c for c in result.columns if c != result.columns[0]]
        resolved_measures = [{"column": measure_col or "", "aggregation": "backlog",
                              "label": "Outstanding"}]

    elif mode == "quality":
        result = profile_completeness(view, df, columns)
        label_columns = ["Column"]
        measure_columns = [c for c in result.columns if c != "Column"]
        resolved_measures = []

    elif mode == "peak":
        resolved_measures = _normalise_measures(measures, view, df)
        if not resolved_measures or not resolved_measures[0].get("column"):
            resolved_measures = [{"column": DEFAULT_MEASURE[view],
                                  "aggregation": "sum",
                                  "label": DEFAULT_MEASURE[view]}]
        group_by = [g for g in (group_by or []) if g]
        for col in group_by:
            _require_column(df, col, "a grouping", view)

        # The position must be built from a comparable measure; adding a local
        # amount across currencies would make the peak meaningless.
        resolved_measures, currency_notes = guard_currency(
            view, df, resolved_measures, group_by)
        target = resolved_measures[0]["column"]

        stamp_col = (time_bucket or {}).get("column") or TIMESTAMP_COLUMN.get(view)
        if not stamp_col:
            raise QueryError(
                f"A peak position needs a timestamp column on the {VIEW_SHEETS[view]}."
            )
        result = compute_peak(df, view, target, stamp_col, group_by,
                              signed_by=SIGN_COLUMN.get(view))
        label_columns = [c for c in result.columns
                         if c in ("Value Date", *group_by)]
        measure_columns = [c for c in result.columns if c not in label_columns]
        # Sorted worst-first, so "the peak" is the head of this table. Showing
        # every day for every currency buries the answer it was asked for.
        total_positions = len(result)
        effective_limit = limit or PEAK_DEFAULT_LIMIT
        if total_positions > effective_limit:
            result = result.head(effective_limit)
            currency_notes.append(
                f"Showing the {effective_limit} largest usage positions of "
                f"{total_positions} day/group combinations, worst first. "
                "Ask for a specific currency or date to see the rest."
            )
        resolved_measures = [{"column": target, "aggregation": "peak",
                              "label": target}]

    elif mode == "distribution":
        # How is a single numeric column distributed? Returns the summary
        # statistics as the table, and carries the raw values so a histogram or
        # box plot can be drawn from them.
        resolved_measures = _normalise_measures(measures, view, df)
        if not resolved_measures or not resolved_measures[0]["column"]:
            raise QueryError(
                "A distribution needs a numeric column. "
                f"Available numeric columns: "
                f"{', '.join(c for c in df.columns if pd.api.types.is_numeric_dtype(df[c]))}."
            )
        group_by = [g for g in (group_by or []) if g]
        for col in group_by:
            _require_column(df, col, "a grouping", view)

        # A distribution mixes currencies just as an aggregation does: the
        # spread of "amount" across JPY and GBP rows describes the exchange
        # rate more than the business.
        resolved_measures, currency_notes = guard_currency(
            view, df, [{**resolved_measures[0], "aggregation": "sum"}], group_by)
        target = resolved_measures[0]["column"]
        df = _as_numeric(df, target)

        if group_by:
            # One distribution per group, compared side by side. Only the
            # comparable statistics: the sigma-coverage columns describe a single
            # distribution's shape and are noise repeated across hundreds of rows.
            compact = ["count", "mean", "median", "std", "min", "max"]
            rows = []
            for key, part in df.groupby(group_by[0], dropna=False):
                stats = describe_series(part[target])
                if stats:
                    rows.append({group_by[0]: str(key),
                                 **{k: stats[k] for k in compact if k in stats}})
            result = pd.DataFrame(rows)
            # Widest spread first: the question behind a grouped distribution is
            # almost always "which of these is most variable".
            if "std" in result.columns:
                result = result.sort_values("std", ascending=False).reset_index(drop=True)
            label_columns = [group_by[0]]
            measure_columns = [c for c in result.columns if c != group_by[0]]
            raw_values = df[[group_by[0], target]].copy()
            if len(result) > MANAGEMENT_ROW_CAP:
                currency_notes.append(
                    f"Showing the {MANAGEMENT_ROW_CAP} most variable of "
                    f"{len(result)} groups, by standard deviation."
                )
                result = result.head(MANAGEMENT_ROW_CAP)
        else:
            stats = describe_series(df[target])
            if not stats:
                raise QueryError(f"Column '{target}' has no numeric values to describe.")
            # One statistic per row reads better than one very wide row.
            result = pd.DataFrame(
                [{"Statistic": k, "Value": v} for k, v in stats.items()]
            )
            label_columns = ["Statistic"]
            measure_columns = ["Value"]
            raw_values = df[[target]].copy()

        resolved_measures = [{"column": target, "aggregation": "distribution",
                              "label": target}]

    elif mode == "list":
        wanted = [c for c in (columns or []) if c]
        for col in wanted:
            _require_column(df, col, "a column", view)
        result = df[wanted].copy() if wanted else df.copy()
        if sort_by:
            _require_column(df, sort_by, "a sort column", view)
            sort_series = df[sort_by]
            if sort_by not in result.columns:
                result = result.assign(**{sort_by: sort_series})
            result = result.sort_values(sort_by, ascending=not sort_desc)
        result = result.reset_index(drop=True)
        if limit:
            result = result.head(limit)
        label_columns = list(result.columns)
        resolved_measures: list[dict[str, str]] = []
    else:
        resolved_measures = _normalise_measures(measures, view, df)
        if not resolved_measures:
            resolved_measures = [
                {"column": DEFAULT_MEASURE[view], "aggregation": "sum",
                 "label": DEFAULT_MEASURE[view]}
            ]

        group_by = [g for g in (group_by or []) if g]
        for col in group_by:
            _require_column(df, col, "a grouping", view)

        resolved_measures, currency_notes = guard_currency(
            view, df, resolved_measures, group_by)

        if time_bucket:
            df, bucket_label = _apply_time_bucket(df, time_bucket, view)
            group_by = [bucket_label] + group_by

        for m in resolved_measures:
            if m["aggregation"] not in NON_NUMERIC_AGGREGATIONS:
                df = _as_numeric(df, m["column"])

        # With no rows and no grouping the honest answer is still a number -
        # "0 transfers failed" is a finding, an empty table is a non-answer.
        if df.empty and group_by:
            result = pd.DataFrame()
        elif group_by:
            frames = []
            for m in resolved_measures:
                if m["aggregation"] == "count":
                    grouped = df.groupby(group_by, dropna=False)
                    part = (grouped[m["column"]].count() if m["column"]
                            else grouped.size()).rename(m["label"])
                else:
                    part = (df.groupby(group_by, dropna=False)[m["column"]]
                            .apply(lambda s, how=m["aggregation"]: _aggregate_series(s, how))
                            .rename(m["label"]))
                frames.append(part)
            result = pd.concat(frames, axis=1).reset_index()
            measure_columns = [m["label"] for m in resolved_measures]
            label_columns = group_by

            # When one measure is split by a small categorical (DR vs CR, say),
            # pivot it into columns. "Amount by sub branch and Dr/Cr mark" then
            # reads as one row per sub branch with a DR and a CR column, which
            # is what a comparison should look like.
            pivoted = False
            pivot_series: list[str] = []
            if (len(group_by) == 2 and len(measure_columns) == 1
                    and result[group_by[1]].nunique() <= 4):
                wide = result.pivot(index=group_by[0], columns=group_by[1],
                                    values=measure_columns[0])
                wide.columns = [str(c) for c in wide.columns]
                result = wide.reset_index().fillna(0)
                measure_columns = [c for c in result.columns if c != group_by[0]]
                label_columns = [group_by[0]]
                pivoted = True
                pivot_series = list(measure_columns)

            if rate:
                result, rate_col = apply_rate(df, result, group_by, rate)
                measure_columns.extend([rate_col, "Matching", "Of total"])
                if not sort_by:
                    sort_by = rate_col

            # Sort by the named column if given, else the first measure.
            # A time series is the exception: chronology IS the axis. Ordering
            # it by size scrambles the chart into meaningless zig-zags and,
            # worse, makes the running total below accumulate out of order, so
            # the cumulative line describes nothing at all.
            bucket_col = None
            if time_bucket:
                _gran = str(time_bucket.get("granularity") or "hour").lower()
                bucket_col = f"{time_bucket.get('column')} ({_gran})"
            if sort_by and sort_by in result.columns:
                order_col, ascending = sort_by, not sort_desc
            elif bucket_col and bucket_col in result.columns:
                order_col, ascending = bucket_col, True
            else:
                order_col, ascending = measure_columns[0], not sort_desc
            result = result.sort_values(order_col, ascending=ascending).reset_index(drop=True)
            if limit:
                result = result.head(limit)
            else:
                result, fold_note = _cap_for_reading(
                    result, label_columns, measure_columns, resolved_measures,
                    is_time_series=bool(time_bucket))
                if fold_note:
                    currency_notes.append(fold_note)
            if pivoted:
                measures_note = f"split by {group_by[1]}"
            else:
                measures_note = None

            # Share of total and running total are the two derived views asked
            # for most often, and both are wrong if computed by hand later.
            primary_measure = measure_columns[0] if measure_columns else None
            if add_share_of_total and primary_measure:
                # A share needs a denominator that means something. Each row may
                # be a valid single-currency total while their sum is not, so
                # the share is computed from the FX-translated column instead.
                share_source, share_label = primary_measure, primary_measure
                ccy_col = CURRENCY_COLUMN.get(view)
                equivalents = DISPLAY_EQUIVALENT.get(view, {})
                mixes_currency = (
                    primary_measure in equivalents
                    and ccy_col in df.columns
                    and df[ccy_col].nunique(dropna=True) > 1
                )
                if mixes_currency:
                    display = equivalents[primary_measure]
                    if not display or display not in df.columns:
                        raise QueryError(
                            f"A share of total cannot be computed from "
                            f"'{primary_measure}': each figure is in its own "
                            f"currency, so the total they are a share of would "
                            f"add unlike units. '{primary_measure}' has no "
                            "FX-translated equivalent in this data."
                        )
                    comparable = (_as_numeric(df, display)
                                  .groupby(group_by, dropna=False)[display].sum())
                    result = result.merge(
                        comparable.rename("__share_base__").reset_index(),
                        on=group_by, how="left")
                    share_source, share_label = "__share_base__", display
                    currency_notes.append(
                        f"Each share is calculated on '{display}', the "
                        f"FX-translated amount, because the currency totals in "
                        f"'{primary_measure}' are in different units and their "
                        "sum would not be a meaningful denominator."
                    )

                values = result[share_source]
                # Debits and credits offset, so a share of the net total can be
                # negative or exceed 100% and describes nothing. Gross flow is
                # the honest denominator when the signs are mixed.
                mixed_sign = bool((values > 0).any() and (values < 0).any())
                total = values.abs().sum() if mixed_sign else values.sum()
                if total:
                    share_name = (f"% of gross {share_label}" if mixed_sign
                                  else f"% of total {share_label}")
                    result[share_name] = (values.abs() / total * 100).round(2) \
                        if mixed_sign else (values / total * 100).round(2)
                    measure_columns.append(share_name)
                    if mixed_sign:
                        currency_notes.append(
                            "Shares are calculated on gross flow because the "
                            "figures include both debits and credits: as a share "
                            "of the net total they would exceed 100% and could be "
                            "negative."
                        )
                if share_source == "__share_base__":
                    result = result.drop(columns="__share_base__")
            if add_cumulative and primary_measure:
                # A running total must not run across the dimensions the table
                # is split by. Summed down a currency-split column it would add
                # JPY to GBP - each grouped figure is valid on its own, and the
                # cumulative silently combines them.
                bucket_label = None
                if time_bucket:
                    gran = str(time_bucket.get("granularity") or "hour").lower()
                    bucket_label = f"{time_bucket.get('column')} ({gran})"
                partition = [c for c in label_columns if c != bucket_label]
                # After a pivot each series is its own column, so accumulating
                # only the first would produce a lone "Cumulative JPY" beside a
                # USD line - a legend entry that answers a question nobody asked.
                targets = pivot_series if pivoted and pivot_series else [primary_measure]
                if partition:
                    for target in targets:
                        result[f"Cumulative {target}"] = (
                            result.groupby(partition, sort=False)[target].cumsum())
                    currency_notes.append(
                        "The running total accumulates within each "
                        + " and ".join(partition)
                        + ", not across them - a total spanning different "
                        "groups would combine values that are not comparable."
                    )
                else:
                    for target in targets:
                        result[f"Cumulative {target}"] = result[target].cumsum()
                measure_columns.extend(f"Cumulative {t}" for t in targets)

            # A time series without a baseline cannot answer "is this normal".
            # Added last so the columns it introduces are never mistaken for
            # grouping dimensions by the calculations above.
            if time_bucket and primary_measure and len(result) >= 5:
                values = result[primary_measure]
                average, spread = values.mean(), values.std(ddof=1)
                if pd.notna(average) and average != 0:
                    result["vs average %"] = ((values - average) / abs(average)
                                              * 100).round(1)
                    measure_columns.append("vs average %")
                if pd.notna(spread) and spread > 0:
                    unusual = (values - average).abs() > 2 * spread
                    if unusual.any():
                        result["Unusual"] = unusual.map({True: "yes", False: ""})
                        currency_notes.append(
                            f"{int(unusual.sum())} of {len(result)} periods sit more "
                            f"than two standard deviations from the period average "
                            f"of {average:,.0f} and are marked as unusual."
                        )
        else:
            row = {}
            for m in resolved_measures:
                if m["aggregation"] == "count":
                    row[m["label"]] = (int(df[m["column"]].count()) if m["column"]
                                       else len(df))
                else:
                    value = _aggregate_series(df[m["column"]], m["aggregation"])
                    row[m["label"]] = value.item() if hasattr(value, "item") else value
            result = pd.DataFrame([row])
            measure_columns = list(result.columns)

    primary = (resolved_measures[0]
               if mode in {"aggregate", "distribution"} and resolved_measures else None)
    return {
        "view": view,
        "sheet": VIEW_SHEETS[view],
        "table": result,
        "mode": mode,
        "label_columns": label_columns,
        "measure_columns": measure_columns,
        "raw_values": raw_values,
        "provenance": {
            "view": view,
            "sheet": VIEW_SHEETS[view],
            "source_file": DATA_FILE.name,
            "mode": mode,
            "filters": filters,
            "group_by": label_columns if mode == "aggregate" else [],
            "time_bucket": time_bucket,
            "measures": ([f"{m['aggregation']} of {m['column'] or 'rows'}"
                          for m in resolved_measures]
                         + ([measures_note] if mode == "aggregate" and measures_note else [])
                         ) if mode in {"aggregate", "distribution"} else [],
            "columns": label_columns if mode == "list" else [],
            "sort_by": sort_by,
            "limit": limit,
            "join": join_info,
            "derived": derived_columns,
            "currency_corrections": currency_notes + ([scope_note] if scope_note else []),
            "data_as_of": data_as_of(view),
            # Kept for existing readers of the provenance block.
            "measure": primary["column"] if primary else None,
            "aggregation": primary["aggregation"] if primary else None,
            "rows_in_view": total_rows,
            "rows_after_filters": rows_after_filters,
            "rows_returned": len(result),
            "executed_at": dt.datetime.now().isoformat(timespec="seconds"),
        },
    }


def declared_row_count(view: str) -> int | None:
    """The row count the view's own control rows claim.

    Each screen prints how many rows it loaded ("Items Loaded", or "Rows"),
    which makes the workbook checkable against itself: if the stated count and
    the loaded count disagree, either the extract was truncated or the loader is
    dropping rows, and both are worth failing over.
    """
    sheet = VIEW_SHEETS.get(view)
    if not sheet:
        return None
    book = openpyxl.load_workbook(DATA_FILE, data_only=True, read_only=True)
    try:
        if sheet not in book.sheetnames:
            return None
        page = book[sheet]
        # The control rows sit between the title and the blank rows above the
        # header, as label/value pairs across the row.
        for row in page.iter_rows(min_row=2, max_row=HEADER_ROW - 1,
                                 values_only=True):
            cells = list(row)
            for index, cell in enumerate(cells[:-1]):
                if str(cell).strip() in {"Items Loaded", "Rows"}:
                    value = cells[index + 1]
                    if isinstance(value, (int, float)):
                        return int(value)
                    if isinstance(value, str) and value.strip().isdigit():
                        return int(value.strip())
    finally:
        book.close()
    return None


def sample_rows(view: str, n: int = 5) -> list[dict[str, Any]]:
    return get_frame(view).head(n).to_dict(orient="records")


# --------------------------------------------------------------------------
# Locating a value the user named
# --------------------------------------------------------------------------
#
# The model is shown column NAMES but, beyond small enums, not their contents.
# So a question that names a thing - "Russia", an account, a venue - gives it
# nothing to match on, and it fails in one of two unrecoverable ways: it filters
# the nearest plausible column ("Counterparty is Russia" matches no rows and is
# reported as an empty view), or it declares the attribute missing when a
# reference table holds it. Both look like answers, so the user cannot tell the
# tool got it wrong.
#
# Indexing the values and telling the model where the ones it was asked about
# actually live turns a guess into a lookup.

# Columns with more distinct values than this are identifiers or free text; a
# question naming one is still worth resolving, but indexing every value of a
# genuinely unbounded column would not pay for itself.
VALUE_INDEX_MAX_DISTINCT = 3000

# The enum cut-off used by schema_summary(): at or below this the model is
# already shown the values, so repeating them as a hint would be noise.
VALUE_VISIBLE_MAX_DISTINCT = 12

# Words common enough in ordinary questions that matching them tells us nothing.
VALUE_STOPWORDS = frozenset({
    "none", "other", "all", "new", "open", "closed", "yes", "no", "true",
    "false", "total", "value", "amount", "date", "time", "type", "name",
    "status", "account", "client", "business", "ledger", "view", "data",
    "local", "display", "group", "cash", "bank", "daily", "month", "day",
})

_TOKEN_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9\-/'\.]*")


@lru_cache(maxsize=1)
def _value_index() -> dict[str, dict[str, Any]]:
    """Map every categorical value in the data to the columns that hold it."""
    index: dict[str, dict[str, Any]] = {}

    view_columns: set[str] = set()
    for _view in VIEW_SHEETS:
        view_columns.update(get_frame(_view).columns)

    def add(value: Any, source: str, column: str, join_on: str | None,
            visible: bool) -> None:
        if not isinstance(value, str):
            return
        text = value.strip()
        # Long strings are prose, not categories worth matching a question on.
        if not text or len(text) > 60:
            return
        entry = index.setdefault(text.casefold(), {"value": text, "places": []})
        # A reference table is only reachable in one hop when its key is a
        # column the transaction data already has. country_master keys on
        # "Country", which no transaction view holds, so it needs a country
        # brought across first - saying so stops the model joining it directly.
        place = {"source": source, "column": column, "join_on": join_on,
                 "visible": visible,
                 "direct": join_on is None or join_on in view_columns}
        if place not in entry["places"]:
            entry["places"].append(place)

    for view in VIEW_SHEETS:
        df = get_frame(view)
        for col in df.columns:
            series = df[col]
            if not pd.api.types.is_string_dtype(series):
                continue
            distinct = series.dropna().unique()
            if len(distinct) > VALUE_INDEX_MAX_DISTINCT:
                continue
            visible = len(distinct) <= VALUE_VISIBLE_MAX_DISTINCT
            for value in distinct:
                add(value, view, col, None, visible)

    for name, df in _reference_frames().items():
        if df.empty:
            continue
        key_column = df.columns[0]
        for col in df.columns:
            distinct = pd.Series(df[col].dropna().unique())
            if len(distinct) > VALUE_INDEX_MAX_DISTINCT:
                continue
            for value in distinct:
                # A reference value is never visible to the model: the prompt
                # lists these tables by column name only.
                add(value, name, col, key_column, False)

    return index


def _token_matches(key: str, tokens: set[str], codes: set[str],
                   bigrams: set[str], question: str) -> bool:
    """Whether an indexed value was named in the question.

    Prefix matching in both directions, because a question says "Russian" and
    the data says "Russia" - an exact match would miss every adjective.

    Short values are codes, and codes collide with ordinary words: "BY" is
    Belarus, "A" is a credit rating, and both appear in almost any sentence.
    They only count when the question wrote them as codes.
    """
    if " " in key:
        # People name things partly: "Zephyr Clearing" for a venue stored as
        # "Zephyr Clearing Hong Kong". Requiring the whole value would miss
        # every abbreviated reference, so two consecutive words are enough.
        return key in question or any(pair in key for pair in bigrams)
    if len(key) <= 3:
        return key.upper() in codes
    if key in tokens:
        return True
    if len(key) < 5:
        return False
    return any(token.startswith(key) or (len(token) >= 5 and key.startswith(token))
               for token in tokens)


def locate_values(question: str, limit: int = 8) -> list[dict[str, Any]]:
    """Values named in the question, and the columns that actually hold them.

    Only values the model cannot already see are returned: small enums are
    listed in the schema it is given, so repeating them would bury the ones
    that matter under things it already knows.
    """
    if not question or not question.strip():
        return []
    index = _value_index()
    lowered = question.casefold()
    raw_tokens = _TOKEN_RE.findall(question)
    tokens = {t.casefold() for t in raw_tokens}
    codes = {t for t in raw_tokens if t.isupper()}
    folded = [t.casefold() for t in raw_tokens]
    bigrams = {f"{a} {b}" for a, b in zip(folded, folded[1:])}

    hits: list[dict[str, Any]] = []
    for key, entry in index.items():
        if key in VALUE_STOPWORDS:
            continue
        if not _token_matches(key, tokens, codes, bigrams, lowered):
            continue
        hidden = [p for p in entry["places"] if not p["visible"]]
        if not hidden:
            continue
        # Only offer places the query engine can actually get to. A reference
        # table's own key is never brought across by a join, and a table whose
        # key no transaction view carries cannot be joined in one hop - naming
        # either sends the model after a column that will not exist.
        reachable = [p for p in hidden if p["direct"]]
        hits.append({
            "value": entry["value"],
            "places": reachable,
            "unreachable": [p for p in hidden if not p["direct"]] if not reachable else [],
            "in_transaction_view": any(p["join_on"] is None
                                       for p in entry["places"]),
        })

    # Longer matches first: they are the more specific thing the user named.
    hits.sort(key=lambda h: (-len(h["value"]), h["value"]))
    return hits[:limit]
