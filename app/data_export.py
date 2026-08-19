"""Workbook and vector exports.

Two formats, for two different jobs:

- **Excel** carries the prepared data plus a native Excel chart. A native chart
  object is the most amendable form for a banking audience: change the range,
  switch the type, restyle it, without any tooling beyond Excel. A pasted image
  would not be amendable at all.
- **SVG** is the chart as vector. It scales without blurring and can be edited
  in a drawing tool or dropped into a deck.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCENT = "8A15E0"
HEADER_FILL = PatternFill("solid", fgColor="F2EAFB")
THIN = Side(style="thin", color="E1E0D9")


def build_workbook(report: dict[str, Any], table: pd.DataFrame, out_dir: Path,
                   session_id: str, stamp: dt.datetime | None = None,
                   measure_columns: list[str] | None = None,
                   label_columns: list[str] | None = None) -> Path:
    """Data, provenance and an editable chart, in one file."""
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = stamp or dt.datetime.now()
    path = out_dir / f"mi_report_{session_id}_{stamp:%Y%m%d_%H%M%S}.xlsx"

    wb = Workbook()
    _sheet_data(wb.active, report, table, measure_columns, label_columns)
    _sheet_about(wb.create_sheet("About this view"), report, stamp)
    wb.save(path)
    return path


def _sheet_data(ws, report: dict[str, Any], table: pd.DataFrame,
                measure_columns: list[str] | None,
                label_columns: list[str] | None) -> None:
    ws.title = "Data"
    ws["A1"] = report.get("title", "MI Report")
    ws["A1"].font = Font(bold=True, size=13, color="0B0B0B")
    ws["A2"] = f'Requested: "{report.get("user_query", "")}"'
    ws["A2"].font = Font(italic=True, size=9, color="52514E")
    ws["A3"] = "Synthetic / non-production data. Figures computed from source, not generated."
    ws["A3"].font = Font(size=8, color="898781")

    start = 5
    for j, col in enumerate(table.columns, start=1):
        cell = ws.cell(row=start, column=j, value=str(col))
        cell.font = Font(bold=True, size=10, color="0B0B0B")
        cell.fill = HEADER_FILL
        cell.border = Border(bottom=Side(style="medium", color=ACCENT))
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(table.iterrows(), start=start + 1):
        for j, col in enumerate(table.columns, start=1):
            value = row[col]
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):
                value = value.item()
            elif isinstance(value, (dt.datetime, dt.date)):
                value = value
            elif not isinstance(value, (int, float, str)):
                value = str(value)
            cell = ws.cell(row=i, column=j, value=value)
            cell.border = Border(bottom=THIN)
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.000"

    for j, col in enumerate(table.columns, start=1):
        longest = max([len(str(col))] + [len(str(v)) for v in table[col].head(60)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, longest + 3), 42)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    _add_native_chart(ws, table, report, start, measure_columns, label_columns)


def _add_native_chart(ws, table: pd.DataFrame, report: dict[str, Any], header_row: int,
                      measure_columns: list[str] | None,
                      label_columns: list[str] | None) -> None:
    """A real Excel chart object, so the reader can amend it in place."""
    if table.empty or len(table.columns) < 2:
        return

    measures = [c for c in (measure_columns or []) if c in table.columns]
    if not measures:
        numeric = [c for c in table.columns if pd.api.types.is_numeric_dtype(table[c])]
        measures = numeric[:3]
    if not measures:
        return
    labels = [c for c in (label_columns or []) if c in table.columns and c not in measures]
    if not labels:
        labels = [c for c in table.columns if c not in measures][:1]
    if not labels:
        return

    n_rows = len(table)
    label_idx = list(table.columns).index(labels[0]) + 1
    chart = LineChart() if report.get("chart_type") == "line" else BarChart()
    if isinstance(chart, BarChart):
        chart.type = "col"
        chart.gapWidth = 60
    chart.title = report.get("title", "MI Report")
    chart.height, chart.width = 9, 20
    chart.style = 2

    for measure in measures[:3]:
        idx = list(table.columns).index(measure) + 1
        data = Reference(ws, min_col=idx, min_row=header_row,
                         max_row=header_row + n_rows)
        chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=label_idx, min_row=header_row + 1,
                     max_row=header_row + n_rows)
    chart.set_categories(cats)

    anchor = f"{get_column_letter(len(table.columns) + 2)}{header_row}"
    ws.add_chart(chart, anchor)


def _sheet_about(ws, report: dict[str, Any], stamp: dt.datetime) -> None:
    """Everything needed to defend the numbers, in the same file as the numbers."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    row = 1

    def section(title: str) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=11, color=ACCENT)
        row += 1

    def pair(key: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, size=9)
        cell = ws.cell(row=row, column=2, value=str(value) if value not in (None, "") else "-")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9)
        row += 1

    def bullets(items: list[str]) -> None:
        nonlocal row
        for item in items or ["- none recorded -"]:
            cell = ws.cell(row=row, column=2, value=f"- {item}")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=9)
            row += 1

    prov = report.get("provenance", {})
    section("What was asked")
    pair("Question", report.get("user_query", ""))
    pair("Interpreted as", report.get("understood", ""))
    pair("Commentary", report.get("narrative", ""))
    row += 1

    section("Limitations")
    bullets(report.get("limitations", []))
    row += 1
    section("Dependencies")
    bullets(report.get("dependencies", []))
    row += 1

    section("How the figures were produced")
    pair("Source", f"{prov.get('sheet', '')} ({prov.get('source_file', '')})")
    pair("Mode", prov.get("mode", ""))
    pair("Measures", ", ".join(prov.get("measures") or []) or "-")
    pair("Grouped by", ", ".join(prov.get("group_by") or []) or "not grouped")
    filters = prov.get("filters") or []
    pair("Filters", "; ".join(
        f"{f.get('column')} {f.get('operator', 'eq')} {f.get('value')}" for f in filters
    ) or "none")
    if prov.get("join"):
        j = prov["join"]
        pair("Combined with", f"{j.get('view')} on {j.get('on')} "
                              f"({j.get('rows_matched')} matched, "
                              f"{j.get('rows_unmatched')} unmatched)")
    if prov.get("derived"):
        pair("Calculated columns", ", ".join(prov["derived"]))
    pair("Rows", f"{prov.get('rows_after_filters', '?')} of {prov.get('rows_in_view', '?')} "
                 f"after filters; {prov.get('rows_returned', '?')} returned")
    pair("Executed", prov.get("executed_at", ""))
    pair("Exported", stamp.strftime("%d %B %Y at %H:%M:%S"))
    row += 1

    section("Important")
    cell = ws.cell(row=row, column=2, value=(
        "Every figure here was computed directly from the source data by a "
        "deterministic query. No figure was generated, estimated or adjusted by a "
        "language model; the AI selected the query and wrote the commentary only. "
        "This is synthetic, non-production data."))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=9, italic=True)


def build_svg(figure_builder, out_dir: Path, session_id: str,
              stamp: dt.datetime | None = None) -> Path | None:
    """Re-render the chart as vector by calling back into the chart builder."""
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = stamp or dt.datetime.now()
    path = out_dir / f"mi_chart_{session_id}_{stamp:%Y%m%d_%H%M%S}.svg"
    return path if figure_builder(path) else None
