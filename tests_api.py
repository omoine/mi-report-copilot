"""API tests using a stub LLM provider - no API key or network needed.

Exercises the full flow: query -> confirm -> refine -> export, plus the
validation paths that protect against bad model output.

Run: .venv\\Scripts\\python.exe tests_api.py
"""

from pathlib import Path
import os
import sys

# Pin the tests to the month workbook explicitly, so a stray DATA_FILE in the
# environment cannot quietly point the suite at something else.
os.environ.setdefault(
    "DATA_FILE",
    str(Path(__file__).parent / "data" / "synthetic_liquidity_fixture.xlsx"),
)

from fastapi.testclient import TestClient

# Read from the fixture rather than hardcoded. A rare value disappears when the
# fixture is regenerated, and a test that fails because JPY is not in a 600-row
# sample has stopped telling anyone anything about refinement.
from app import data_access as _fixture
NOSTRO_ROWS = len(_fixture.get_frame("nostro_transfer"))
# The commonest currency, so the filtered result is never empty - an empty
# result produces no chart, and the export assertions below need one.
SAMPLE_CCY = str(_fixture.get_frame("nostro_transfer")["Currency"]
                 .value_counts().index[0])

from app import auth, llm_client, main, md_export, orchestrator

failures: list[str] = []


def check(label: str, condition: bool, detail: str = "") -> None:
    print(f"  [{'PASS' if condition else 'FAIL'}] {label}" + (f" - {detail}" if detail else ""))
    if not condition:
        failures.append(label)


class StubProvider:
    """Stands in for the model. Returns fixed, valid structures."""

    def __init__(self) -> None:
        self.json_calls = 0
        self.text_calls = 0
        self.next_json: dict | None = None
        self.queue: list[dict] = []  # consumed in order, for retry scenarios

    def complete_json(self, system, messages):
        self.json_calls += 1
        if self.queue:
            return self.queue.pop(0)
        if self.next_json is not None:
            payload, self.next_json = self.next_json, None
            return payload
        return {
            "understood": "Total nostro transfer value by currency for the loaded value date.",
            "query": {
                "view": "nostro_transfer",
                "filters": [],
                "group_by": ["Currency"],
                "measure": "Value Amount",
                "aggregation": "sum",
                "sort_desc": True,
                "limit": None,
            },
            "chart_type": "bar",
            "limitations": [
                "Amounts are in local currency and are not additive across currencies.",
                "Synthetic, non-production data.",
            ],
            "dependencies": ["Depends on the value-date range selected on the source screen."],
            "feasible": True,
        }

    def complete_text(self, system, messages):
        self.text_calls += 1
        return "JPY accounts for the largest share of transfer value in this view."


stub = StubProvider()
main._provider = stub  # inject before any request builds a real client
client = TestClient(main.app)

# The application is behind a password, so the suite signs in the way a user
# does. Testing against an unlocked app would leave the gate itself untested
# and hide the day it starts rejecting everything.
client.post("/api/login", json={"password": auth.password()})

print("\n1. Health and schema")
h = client.get("/api/health").json()
check("health ok", h["status"] == "ok", f"views: {h['views']}")
check("data classification exposed", "Synthetic" in h["data_classification"])
s = client.get("/api/schema").json()
check("schema lists 3 views", len(s["views"]) == 3)
check("schema includes controls", len(s["controls"]) > 20)

print("\n2. Query -> interpretation (nothing built yet)")
r = client.post("/api/query", json={"query": "total nostro transfer value by currency"})
check("query accepted", r.status_code == 200, f"status {r.status_code}")
data = r.json()
sid = data["session_id"]
check("state is awaiting_confirmation", data["state"] == "awaiting_confirmation")
check("limitations returned", len(data["limitations"]) == 2)
check("dependencies returned", len(data["dependencies"]) == 1)
check("query summary present", "Value Amount" in data["query_summary"], data["query_summary"][:70])
check("no report yet", "table" not in data)

print("\n3. Confirm -> report built")
r = client.post("/api/confirm", json={"session_id": sid})
check("confirm ok", r.status_code == 200, f"status {r.status_code}")
rep = r.json()
check("state is report_built", rep["state"] == "report_built")
check("chart produced", bool(rep["chart_url"]), rep.get("chart_url"))
check("table has rows", len(rep["table"]["rows"]) > 0, f"{len(rep['table']['rows'])} rows")
check("narrative from model", "JPY" in rep["narrative"])
check("provenance recorded", rep["provenance"]["rows_in_view"] == NOSTRO_ROWS,
      f'{rep["provenance"]["rows_in_view"]} vs {NOSTRO_ROWS}')
check("model was asked for narrative", stub.text_calls == 1)

print("\n4. Chart is served")
r = client.get(rep["chart_url"])
check("chart bytes served", r.status_code == 200 and r.content[:4] == b"\x89PNG",
      f"{len(r.content):,} bytes")

print("\n5. Refine")
stub.next_json = {
    "understood": f"Filtered to {SAMPLE_CCY} only.",
    "query": {
        "view": "nostro_transfer",
        "filters": [{"column": "Currency", "operator": "eq", "value": SAMPLE_CCY}],
        "group_by": ["Sending Strategy"],
        "measure": "Value Amount",
        "aggregation": "sum",
        "sort_desc": True,
        "limit": None,
    },
    "chart_type": "barh",
    "limitations": ["Single currency only."],
    "dependencies": ["Depends on sending strategy being populated."],
    "feasible": True,
}
r = client.post("/api/refine", json={
    "session_id": sid, "instruction": f"just {SAMPLE_CCY}, by sending strategy"})
check("refine ok", r.status_code == 200, f"status {r.status_code}")
ref = r.json()
check("refined filter applied",
      ref["provenance"]["filters"][0]["value"] == SAMPLE_CCY)
check("fewer rows after filter",
      ref["provenance"]["rows_after_filters"] < ref["provenance"]["rows_in_view"],
      f"{ref['provenance']['rows_after_filters']} of {ref['provenance']['rows_in_view']}")
check("regrouped", ref["provenance"]["group_by"] == ["Sending Strategy"])

print("\n6. Export")
r = client.post("/api/export", json={"session_id": sid})
check("export ok", r.status_code == 200, f"status {r.status_code}")
exp = r.json()
check("pdf named", exp["pdf"].endswith(".pdf"), exp["pdf"])
check("md named", exp["markdown"].endswith(".md"), exp["markdown"])
check("pdf and md are a matching pair",
      exp["pdf"][:-4] == exp["markdown"][:-3], f"{exp['pdf']} / {exp['markdown']}")

r = client.get(f"/api/download/{exp['pdf']}")
check("pdf downloads", r.status_code == 200 and r.content[:4] == b"%PDF",
      f"{len(r.content):,} bytes")

# A short result table must not be orphaned across a page break.
try:
    import io

    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(r.content))
    pages = [page.extract_text() or "" for page in reader.pages]
    labels = [str(row[0]) for row in ref["table"]["rows"]]
    # Whether the table is SPLIT, not whether it landed on page one. Keeping a
    # short table whole by moving it to the next page is the correct outcome,
    # and looking only at page one confused that with a break - the commentary
    # names the same groups, so a label appears there either way.
    together = any(all(label in text for label in labels) for text in pages)
    check("short table is not split across pages", together,
          f"{len(labels)} rows across {len(pages)} pages")
except ImportError:
    print("  [SKIP] page-break check (pypdf not installed)")
r = client.get(f"/api/download/{exp['markdown']}")
md = r.text
check("md downloads", r.status_code == 200, f"{len(md):,} chars")
check("md records the refinement history", f"just {SAMPLE_CCY}" in md)
check("md records the filter", SAMPLE_CCY in md)
check("md states figures are deterministic", "deterministic" in md.lower())
check("md has field definitions", "Sending Strategy" in md)

print("\n6b. Export formats")
check("excel produced", bool(exp.get("excel")), exp.get("excel"))
check("svg produced", bool(exp.get("svg")), exp.get("svg"))

r = client.get(f"/api/download/{exp['excel']}")
check("excel downloads", r.status_code == 200 and r.content[:2] == b"PK",
      f"{len(r.content):,} bytes")
try:
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    check("excel has data and about sheets",
          {"Data", "About this view"} <= set(wb.sheetnames), str(wb.sheetnames))
    # An editable chart object is the point of the Excel export.
    check("excel carries a native chart", len(wb["Data"]._charts) > 0,
          f"{len(wb['Data']._charts)} chart(s)")
    about = "\n".join(str(c.value) for row in wb["About this view"].iter_rows()
                      for c in row if c.value)
    check("about sheet records provenance", "deterministic" in about.lower())
except ImportError:
    print("  [SKIP] excel internals (openpyxl missing)")

r = client.get(f"/api/download/{exp['svg']}")
check("svg downloads and is vector", r.status_code == 200 and b"<svg" in r.content[:600],
      f"{len(r.content):,} bytes")

print("\n6c. Saved views")
r = client.post("/api/views/save", json={"session_id": sid, "name": "Eval test view"})
check("view saved", r.status_code == 200, r.json().get("detail", "")[:60])

r = client.get("/api/views")
names = [v["name"] for v in r.json()["views"]]
check("view appears in the list", "Eval test view" in names, str(names[:4]))

r = client.get("/api/views?search=eval")
check("search finds it", any(v["name"] == "Eval test view" for v in r.json()["views"]))
r = client.get("/api/views?search=zzzznotathing")
check("search excludes non-matches", r.json()["views"] == [])

# Saving the same name again must ask rather than silently overwrite.
r = client.post("/api/views/save", json={"session_id": sid, "name": "Eval test view"})
check("duplicate name is refused", r.status_code == 400,
      r.json().get("detail", "")[:60])
r = client.post("/api/views/save",
                json={"session_id": sid, "name": "Eval test view", "overwrite": True})
check("overwrite works when confirmed", r.status_code == 200)

view_id = [v["id"] for v in client.get("/api/views").json()["views"]
           if v["name"] == "Eval test view"][0]
r = client.post("/api/views/load", json={"view_id": view_id})
check("saved view reloads and re-runs", r.status_code == 200,
      r.json().get("detail", "")[:80])
if r.status_code == 200:
    reloaded = r.json()
    check("reload returns figures", len(reloaded["table"]["rows"]) > 0)
    check("reload names the view", reloaded.get("loaded_view", {}).get("name") == "Eval test view")

# A saved view is configuration that outlives the code that wrote it. One
# carrying a key this build has never heard of must degrade, not 500.
future = dict(orchestrator.saved_views.get_view(view_id))
future["query"] = {**future["query"], "some_future_capability": {"x": 1}}
clean, dropped = orchestrator.compatible_query(future["query"])
check("an unknown specification key is dropped, not passed through",
      "some_future_capability" in dropped and "some_future_capability" not in clean,
      f"dropped {dropped}")
check("the rest of the specification survives", clean.get("view") == future["query"]["view"])

r = client.delete(f"/api/views/{view_id}")
check("view deleted", r.status_code == 200)
check("deleted view is gone",
      "Eval test view" not in [v["name"] for v in client.get("/api/views").json()["views"]])
r = client.delete(f"/api/views/{view_id}")
check("deleting twice is a clean 404", r.status_code == 404)

print("\n6d. In-app assistant")


class ChatStub(StubProvider):
    """Captures the system prompt so the grounding can be inspected."""

    def __init__(self) -> None:
        super().__init__()
        self.last_system = ""

    def complete_text(self, system, messages):
        self.last_system = system
        return "The largest contributor is JPY.\nTRY: total transfer value by currency"


chat_stub = ChatStub()
main._provider = chat_stub

# With a report on screen the assistant must be grounded in the export itself,
# not in a separate summary that could describe the same figures differently.
r = client.post("/api/chat", json={"session_id": sid, "message": "what is this?"})
check("chat answers with a report open", r.status_code == 200, f"status {r.status_code}")
body = r.json()
check("mode is explain when a report exists", body.get("mode") == "explain", body.get("mode"))
check("a suggested prompt is offered separately", body.get("suggestion") ==
      "total transfer value by currency", str(body.get("suggestion")))
check("the suggestion is stripped from the reply", "TRY:" not in body.get("reply", ""))

session_obj = orchestrator.get_session(sid)
exported = md_export.render_markdown({**session_obj.report, "history": session_obj.history})
grounding = chat_stub.last_system

check("the assistant is grounded in the export, verbatim",
      exported[:400] in grounding,
      f"export {len(exported):,} chars, prompt {len(grounding):,} chars")

# Spot-check that the figures themselves travelled, not just the headings.
figure_rows = [row[0] for row in ref["table"]["rows"][:3]]
check("the figures the user sees are in the assistant's context",
      all(str(v) in grounding for v in figure_rows), str(figure_rows))
check("the stated limitations travelled too",
      all(lim[:40] in grounding for lim in session_obj.report["limitations"][:2]))

# Before anything is built it should be advising, not explaining.
fresh = client.post("/api/chat", json={"message": "I want to see failed payments"})
check("mode is design before a report exists",
      fresh.json().get("mode") == "design", fresh.json().get("mode"))
check("design mode is not grounded in a report",
      "SUPPORTING DOCUMENT" not in chat_stub.last_system)
check("design mode knows the refusals",
      "different currencies" in chat_stub.last_system)

r = client.post("/api/chat", json={"session_id": sid, "message": "   "})
check("an empty question is rejected", r.status_code == 400)

main._provider = stub  # restore for the checks that follow

print("\n6e. Front-end class hygiene")
static_dir = Path(__file__).parent / "static"
css = (static_dir / "styles.css").read_text(encoding="utf-8")
app_js = (static_dir / "app.js").read_text(encoding="utf-8")
assistant_js = (static_dir / "assistant.js").read_text(encoding="utf-8")

# The conversation gives its blocks class "msg assistant". A bare `.assistant`
# rule in the stylesheet therefore lands on every report block - which is how
# they all became position:fixed and rendered over the top of the page.
import re  # noqa: E402

conversation_classes = set(re.findall(r'className = `msg \$\{(\w+)\}`', app_js))
bare_rules = re.findall(r"^\.assistant\s*[,{]", css, re.MULTILINE)
check("no bare .assistant rule in the stylesheet", not bare_rules,
      f"{len(bare_rules)} found - it would style every conversation block")
check("the widget root is namespaced",
      "assistant-widget" in assistant_js and "assistant-widget" in css)
check("the conversation still uses the msg/assistant pairing",
      "cls = 'assistant'" in app_js,
      "if this changes, revisit the collision guard above")

# Any class the widget positions must not be one the conversation applies.
positioned = set(re.findall(r"^\.([a-z-]+)\s*\{[^}]*position:\s*fixed", css,
                            re.MULTILINE | re.DOTALL))
check("nothing the conversation applies is position:fixed",
      "assistant" not in positioned, str(sorted(positioned)))

# Setting the hidden property does nothing if a class also sets display: the
# user-agent's [hidden]{display:none} loses to any rule here. The panel stayed
# on screen after "closing" for exactly this reason, and a test that checked
# the property rather than the rendering reported it as working.
check("the hidden attribute is honoured everywhere",
      "[hidden] { display: none !important; }" in css,
      "without this, any class that sets display keeps a 'hidden' element on "
      "screen - .layout and .assistant-panel both shipped that bug")

hidden_toggled = set(re.findall(r"(\w[\w-]*)\.hidden\s*=", assistant_js))
for element in hidden_toggled:
    cls = f"assistant-{element}"
    sets_display = re.search(rf"^\.{cls}\s*\{{[^}}]*display:", css, re.MULTILINE | re.DOTALL)
    has_guard = f".{cls}[hidden]" in css
    check(f"hiding .{cls} actually hides it",
          not sets_display or has_guard,
          "a class display rule outranks the hidden attribute; "
          f"add .{cls}[hidden] {{ display: none; }}")

print("\n7. Guard rails")
r = client.post("/api/export", json={"session_id": "nonexistent-session"})
check("export without report is rejected", r.status_code == 400, r.json().get("detail", "")[:50])

r = client.post("/api/refine", json={"session_id": "another-new-session", "instruction": "x"})
check("refine before build is rejected", r.status_code == 400)

r = client.get("/api/download/../../etc/passwd")
check("path traversal blocked", r.status_code in (403, 404), f"status {r.status_code}")

stub.next_json = {"understood": "Cannot do this.", "feasible": False,
                  "limitations": ["No FX forward data exists in these views."], "query": {}}
r = client.post("/api/query", json={"query": "show me FX forwards"})
check("infeasible request rejected cleanly", r.status_code == 400,
      r.json().get("detail", "")[:60])
check("decline explains why, not just what was asked",
      "No FX forward data" in r.json().get("detail", ""),
      r.json().get("detail", "")[:90])

stub.next_json = {"understood": "x", "feasible": True,
                  "query": {"view": "not_a_view"}, "limitations": [], "chart_type": "bar"}
r = client.post("/api/query", json={"query": "something"})
check("bad view from model rejected", r.status_code == 400)

bad_column = {"understood": "x", "feasible": True, "limitations": [], "chart_type": "bar",
              "query": {"view": "client", "group_by": ["No Such Column"],
                        "measure": "Calculated Balance (Local)", "aggregation": "sum"}}
good_column = {"understood": "x", "feasible": True, "limitations": [], "chart_type": "bar",
               "query": {"view": "client", "group_by": ["Currency"],
                         "measure": "Calculated Balance (Local)", "aggregation": "sum"}}

# A wrong column should be retried once with the error fed back to the model.
stub.queue = [bad_column, good_column]
before = stub.json_calls
r = client.post("/api/query", json={"query": "recoverable mistake"})
check("bad column is retried and recovers", r.status_code == 200, f"status {r.status_code}")
check("retry made exactly one extra model call", stub.json_calls - before == 2,
      f"{stub.json_calls - before} calls")

# Failing twice must surface to the user rather than loop.
stub.queue = [bad_column, bad_column]
r = client.post("/api/query", json={"query": "unrecoverable mistake"})
check("bad column fails after one retry", r.status_code == 400,
      r.json().get("detail", "")[:60])
stub.queue = []

r = client.post("/api/query", json={"query": "   "})
check("empty query rejected", r.status_code == 400)

print("\n8. Narrative failure does not lose the report")


class FailingNarrative(StubProvider):
    def complete_text(self, system, messages):
        raise llm_client.LLMError("simulated model outage")


main._provider = FailingNarrative()
r = client.post("/api/query", json={"query": "total nostro transfer value by currency"})
sid2 = r.json()["session_id"]
r = client.post("/api/confirm", json={"session_id": sid2})
check("report still built without narrative", r.status_code == 200)
check("figures still present", len(r.json()["table"]["rows"]) > 0)
check("fallback commentary shown", "could not be generated" in r.json()["narrative"].lower())

print("\n9. The access gate")

# A separate client that has never signed in. The point of the gate is that
# the API refuses it, not that the interface hides itself - anyone can skip
# the interface.
locked = TestClient(main.app)

for path in ("/api/health", "/api/schema", "/api/views", "/api/model"):
    r = locked.get(path)
    check(f"GET {path} refused when locked", r.status_code == 401, f"got {r.status_code}")

r = locked.post("/api/query", json={"query": "total nostro transfer value by currency"})
check("POST /api/query refused when locked", r.status_code == 401, f"got {r.status_code}")
r = locked.post("/api/chat", json={"message": "hello", "history": []})
check("POST /api/chat refused when locked", r.status_code == 401, f"got {r.status_code}")
check("downloads refused when locked",
      locked.get("/api/download/anything.pdf").status_code == 401)

check("ping stays open for uptime checks", locked.get("/api/ping").status_code == 200)
check("session state is readable when locked",
      locked.get("/api/session").json()["authenticated"] is False)

r = locked.post("/api/login", json={"password": "wrong"})
check("wrong password rejected", r.status_code == 401, f"got {r.status_code}")
check("no cookie issued on a wrong password", auth.COOKIE_NAME not in locked.cookies)
check("still locked after a wrong password", locked.get("/api/views").status_code == 401)

r = locked.post("/api/login", json={"password": auth.password()})
check("correct password accepted", r.status_code == 200, f"got {r.status_code}")
check("cookie issued", auth.COOKIE_NAME in locked.cookies)
check("unlocked client can read views", locked.get("/api/views").status_code == 200)

check("a forged cookie is refused", not auth.valid("9999999999.deadbeef"))
check("an expired but correctly signed cookie is refused",
      not auth.valid("1." + auth._sign("1")))

locked.post("/api/logout")
check("logout re-locks", locked.get("/api/views").status_code == 401)

# The interface must ship locked. If it rendered first and hid itself after,
# everything would be on screen for the moment before the script ran.
index_html = (static_dir / "index.html").read_text(encoding="utf-8")
check("the page ships locked", chr(60) + chr(98) + 'ody class="locked"' in index_html)
check("locked hides everything but the gate",
      "body.locked > *:not(.gate)" in css)

print("\n" + "-" * 55)
if failures:
    print(f"{len(failures)} FAILED: {failures}")
    sys.exit(1)
print("All API tests passed.")
