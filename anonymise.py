"""Deterministic anonymisation with a persistent golden source.

The mapping lives in `anonymisation/mapping.json`, which is git-ignored: it is
the only place real values are ever written, and it must stay in the working
folder. This script contains no real values.

Why a persistent mapping rather than a hash or a fresh pass each time: the same
real value must become the same surrogate in every extract, forever. Otherwise
account 12345 is "Account A" in one file and "Account C" in the next, and no
analysis spanning two extracts is possible.

Usage
-----
Scan a new file and report what is not yet mapped (changes nothing):
    python anonymise.py scan <file.xlsx>

Anonymise a file, extending the mapping with any new values:
    python anonymise.py apply <file.xlsx> --out <anonymised.xlsx>

Show the mapping as a readable table:
    python anonymise.py report [--category counterparty]

Check a file for values that look like they escaped anonymisation:
    python anonymise.py audit <file.xlsx>
"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent
GOLDEN_DIR = ROOT / "anonymisation"
MAPPING_FILE = GOLDEN_DIR / "mapping.json"

# ---------------------------------------------------------------------------
# Which columns carry which kind of name. Extend this when new columns appear -
# a column that is not listed here is passed through untouched, so anything
# sensitive must be registered.
# ---------------------------------------------------------------------------
COLUMN_CATEGORIES: dict[str, str] = {
    # identifiers
    "Source Account": "account_number",
    "Target Account": "account_number",
    "Account": "account_number",
    "Sort Code": "sort_code",
    "Account Name": "account_name",
    "Ledger Account": "ledger_account",
    "Book Id": "book_id",
    # organisations and places
    "Legal Entity": "legal_entity",
    "Counterparty": "counterparty",
    "Source Account Venue Location": "venue",
    "Target Account Venue Location": "venue",
    # people
    "Created By": "person",
    "Approved By": "person",
    "Requested By": "person",
    "Modified By": "person",
    # systems and internal codes - the ones most often forgotten
    "Sending Strategy": "system",
    "Message Status": "system_status",
    "Sub Branch": "branch_code",
    "Cashflow Type": "type_code",
    "Net Type": "type_code",
    "Trade Status": "status_code",
    "Transfer Status": "status_code",
    # references
    "Reference": "reference",
    "Transaction Reference": "reference",
    "Upstream Transaction ID": "reference",
    # groupings and teams - these name desks and org units
    "Group": "org_unit",
    "Client Group": "org_unit",
    "Business Group": "org_unit",
    "Team": "org_unit",
    "Desk": "org_unit",
    "Portfolio": "org_unit",
    # free text can quote anything
    "Comment": "free_text",
    "Narrative": "free_text",
    "Description": "free_text",
}

# Sheets that describe the data rather than containing it. They still carry
# names (the View Controls sheet holds group and team names), so they are
# reported for manual review instead of being silently passed over.
METADATA_SHEETS = {"README", "Data Dictionary", "View Controls", "Reference Data"}

# Table footers and row-count captions that are not real values.
ARTEFACTS = re.compile(r"^(total|subtotal|grand total|n/a|-|\d+\s+records?)$", re.I)

# Values that are genuinely public standards and must NOT be renamed, because
# renaming them would destroy meaning. Everything else in a mapped column is
# treated as potentially internal.
PUBLIC_TERMS = {
    # industry networks and message standards
    "SWIFT", "ISO", "ISO20022", "MT103", "MT202", "SEPA", "CHAPS", "TARGET2",
    "FEDWIRE", "CHIPS", "BACS", "RTGS", "CLS", "EBA", "SIC",
    # currencies
    "GBP", "EUR", "USD", "JPY", "CHF", "HKD", "SGD", "AUD", "NOK", "SEK",
    "ZAR", "CNY", "INR", "CAD", "NZD", "DKK", "PLN", "MXN",
    # generic banking vocabulary and outcome words. Kept legible on purpose:
    # renaming these buys no privacy and makes the anonymised set unreadable
    # ("LEDGER_RECEIVED" turning into "SYWDWR_RECEIVED" helps nobody).
    "NOSTRO", "VOSTRO", "LEDGER", "CASHFLOW", "SETTLEMENT", "TREASURY",
    "LIQUIDITY", "PAYMENT", "TRANSFER", "BALANCE", "FUNDING", "INTERNAL",
    "EXTERNAL", "GROSS", "NET", "DR", "CR", "DEBIT", "CREDIT",
    "NEW", "AMEND", "CANCEL", "PENDING", "APPROVAL", "APPROVED", "RECEIVED",
    "ACKNOWLEDGED", "REJECTED", "FAILED", "SETTLED", "QUEUED", "RELEASED",
    "OR", "AND", "ALL", "BOTH", "NONE", "TOTAL",
}

# Surrogate pools. Deliberately bland and obviously invented.
ORG_STEMS = ["Alder", "Bramble", "Cedar", "Dunlin", "Elmwood", "Fernway",
             "Granite", "Harrow", "Ironbark", "Juniper", "Kestrel", "Larkspur",
             "Mallow", "Nightjar", "Oriel", "Pinehurst", "Quillon", "Redstart",
             "Sablewood", "Thornbury", "Umberton", "Vetchling", "Wrenfield",
             "Yarrow", "Zephyr"]
ORG_SUFFIX = ["Bank", "Trust", "Financial", "Markets", "Securities", "Capital",
              "Treasury Services", "Commercial Bank", "Clearing"]
ORG_LEGAL = ["plc", "N.V.", "S.A.", "Ltd", "AG", "N.A.", "GmbH", "Pte Ltd"]
CITIES = ["Northport", "Eastcliff", "Westmere", "Southgate", "Fairhaven",
          "Kingsbridge", "Millbrook", "Oakfield", "Riverton", "Stonebury"]
FIRST_NAMES = ["Avery", "Blake", "Casey", "Devon", "Emery", "Finley", "Greer",
               "Harper", "Indigo", "Jules", "Kai", "Lennox", "Marlow", "Noor",
               "Oakley", "Payton", "Quinn", "Reese", "Sawyer", "Tatum"]
LAST_NAMES = ["Ashford", "Brookes", "Calloway", "Delaney", "Ellison", "Fairweather",
              "Gallagher", "Hollis", "Isley", "Jennings", "Kirby", "Lindqvist",
              "Merrick", "Novak", "Osgood", "Prentice", "Rhodes", "Sinclair"]


def _category_for(column: str) -> str | None:
    """Look up a column's category, tolerating the __n suffix added to
    duplicate headers."""
    if column in COLUMN_CATEGORIES:
        return COLUMN_CATEGORIES[column]
    base = column.split("__")[0]
    return COLUMN_CATEGORIES.get(base)


def _load() -> dict[str, Any]:
    if MAPPING_FILE.exists() and MAPPING_FILE.stat().st_size:
        return json.loads(MAPPING_FILE.read_text(encoding="utf-8"))
    return {"_meta": {"note": "Golden source. Real values. Never commit."},
            "categories": {}}


def _save(mapping: dict[str, Any]) -> None:
    GOLDEN_DIR.mkdir(exist_ok=True)
    MAPPING_FILE.write_text(json.dumps(mapping, indent=2, ensure_ascii=False),
                            encoding="utf-8")


def _used(mapping: dict[str, Any], category: str) -> set[str]:
    return set(mapping["categories"].get(category, {}).values())


def _surrogate(category: str, real: str, mapping: dict[str, Any]) -> str:
    """Invent a surrogate that preserves the shape of the original.

    Shape matters: an account number that stops being 13 digits, or a code that
    stops being 4 characters, can break downstream parsing and makes the
    anonymised set behave differently from the real one.
    """
    taken = _used(mapping, category)
    rnd = random.Random(f"{category}:{real}")  # stable for a given input

    def unique(make) -> str:
        for _ in range(4000):
            candidate = make()
            if candidate not in taken:
                return candidate
        return f"{make()}-{rnd.randint(1000, 9999)}"

    if category in {"account_number", "sort_code"}:
        digits = len(re.sub(r"\D", "", real)) or 10
        return unique(lambda: "".join(str(rnd.randint(0, 9)) for _ in range(digits)))

    if category == "person":
        return unique(lambda: f"{rnd.choice(FIRST_NAMES)} {rnd.choice(LAST_NAMES)}")

    if category == "counterparty":
        return unique(lambda: f"{rnd.choice(ORG_STEMS)} {rnd.choice(ORG_SUFFIX)}")

    if category == "legal_entity":
        return unique(lambda: f"{rnd.choice(ORG_STEMS)} Group {rnd.choice(ORG_LEGAL)}")

    if category == "venue":
        return unique(lambda: f"{rnd.choice(ORG_STEMS)} {rnd.choice(ORG_SUFFIX)} "
                              f"{rnd.choice(CITIES)}")

    if category in {"branch_code", "book_id", "type_code", "system"}:
        # Keep the original length so column widths and parsers still work.
        n = max(2, min(len(real.strip()), 8))
        prefix = {"branch_code": "BR", "book_id": "BK",
                  "type_code": "TC", "system": "SY"}[category]
        body = n - len(prefix)
        if body <= 0:
            return unique(lambda: "".join(rnd.choice("ABCDEFGHJKLMNPQRSTUVWXYZ")
                                          for _ in range(n)))
        return unique(lambda: prefix + "".join(
            rnd.choice("ABCDEFGHJKLMNPQRSTUVWXYZ") for _ in range(body)))

    if category in {"status_code", "system_status"}:
        # Statuses often embed a system name (<SYSTEM>_REJECTED). Keep the
        # outcome word, replace the system part.
        parts = real.split("_")
        if len(parts) > 1:
            head = _map_value(parts[0], "system", mapping)
            return "_".join([head] + parts[1:])
        return unique(lambda: f"ST{rnd.randint(10, 99)}")

    if category == "ledger_account":
        # Structured like PREFIX.NOSTRO.CHF.<DESK>.003 - keep the structure,
        # remap only the parts that name something internal.
        parts = real.split(".")
        out = []
        for part in parts:
            if part.upper() in PUBLIC_TERMS or part.isdigit():
                out.append(part)
            else:
                out.append(_map_value(part, "branch_code", mapping))
        return ".".join(out)

    if category == "account_name":
        return unique(lambda: f"Account {rnd.choice(ORG_STEMS)} "
                              f"{rnd.randint(1, 99):02d}")

    if category == "reference":
        # Preserve the letter/digit shape so format checks still pass.
        shape = re.sub(r"[A-Za-z]", "A", re.sub(r"\d", "9", real))
        return unique(lambda: "".join(
            rnd.choice("ABCDEFGHJKLMNPQRSTUVWXYZ") if c == "A"
            else str(rnd.randint(0, 9)) if c == "9" else c for c in shape))

    if category == "free_text":
        return "[redacted free text]"

    return unique(lambda: f"VALUE_{rnd.randint(1000, 9999)}")


def _map_value(real: Any, category: str, mapping: dict[str, Any]) -> str:
    """Return the surrogate for one value, creating it on first sight."""
    text = str(real).strip()
    if not text or text.lower() in {"nan", "none", "nat"} or ARTEFACTS.match(text):
        return text
    if text.upper() in PUBLIC_TERMS:
        return text  # public standard, keep it meaningful

    table = mapping["categories"].setdefault(category, {})
    if text not in table:
        table[text] = _surrogate(category, text, mapping)
    return table[text]


def _sweep_known(df: pd.DataFrame, mapping: dict[str, Any]) -> int:
    """Replace any cell whose exact value is already mapped.

    Used on metadata sheets, where there are no columns to key off. Only
    substitutes values already established elsewhere, so it cannot invent a
    surrogate for something that was never seen in the data itself.
    """
    lookup: dict[str, str] = {}
    for table in mapping["categories"].values():
        for real, fake in table.items():
            lookup[real.casefold()] = fake

    replaced = 0
    for col in df.columns:
        series = df[col]
        # Not `== object`: pandas 3 gives text columns a dedicated StringDtype,
        # so an object check silently skips every string column.
        if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
            continue
        new_values = []
        for value in series:
            text = str(value).strip() if value is not None else ""
            hit = lookup.get(text.casefold())
            if hit is not None and text:
                new_values.append(hit)
                replaced += 1
            else:
                new_values.append(value)
        df[col] = new_values
    return replaced


def _read(path: Path) -> dict[str, pd.DataFrame]:
    if path.suffix.lower() in {".csv", ".tsv"}:
        sep = "\t" if path.suffix.lower() == ".tsv" else ","
        return {"data": pd.read_csv(path, sep=sep, dtype=str)}
    wb = openpyxl.load_workbook(path, data_only=True)
    frames = {}
    try:
        for name in wb.sheetnames:
            rows = list(wb[name].iter_rows(values_only=True))
            if not rows:
                continue
            header_idx = _find_header(rows)
            if header_idx is None:
                continue
            header = [str(c).strip() if c is not None else f"col{i}"
                      for i, c in enumerate(rows[header_idx])]
            frames[name] = pd.DataFrame(rows[header_idx + 1:],
                                        columns=_unique_headers(header))
    finally:
        wb.close()
    return frames


def _unique_headers(header: list[str]) -> list[str]:
    """Disambiguate repeated column names, which real extracts often contain.

    Without this, df[col] returns a DataFrame rather than a Series and every
    per-column operation breaks. The suffix keeps the original name as the
    prefix so category lookup still matches on the first occurrence.
    """
    seen: dict[str, int] = {}
    out: list[str] = []
    for name in header:
        if name in seen:
            seen[name] += 1
            out.append(f"{name}__{seen[name]}")
        else:
            seen[name] = 0
            out.append(name)
    return out


def _find_header(rows: list) -> int | None:
    """Pick the row that best looks like a header.

    Scored rather than first-match: control rows above the table often contain
    one or two words that are also column names ("Group", "Team"), so the first
    row with a couple of hits is not reliably the header. The real header has
    many recognised names and few blanks.
    """
    best, best_score = None, 0
    for i, row in enumerate(rows[:15]):
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(cells) < 3:
            continue
        known = sum(1 for c in cells if _category_for(c))
        # Density matters: a header fills most of its row, a control row does not.
        density = len(cells) / max(len(row), 1)
        score = known * 2 + len(cells) * density
        if score > best_score:
            best, best_score = i, score
    if best is not None:
        return best
    for i, row in enumerate(rows[:15]):
        if sum(1 for c in row if c is not None) >= 4:
            return i
    return None


def cmd_scan(path: Path) -> int:
    mapping = _load()
    frames = _read(path)
    unmapped: dict[str, set] = {}
    unregistered: set[str] = set()

    metadata_seen = []
    for sheet, df in frames.items():
        if sheet in METADATA_SHEETS:
            metadata_seen.append(sheet)
            continue
        for col in df.columns:
            category = _category_for(col)
            if category is None:
                series = df[col]
                is_text = not (pd.api.types.is_numeric_dtype(series)
                               or pd.api.types.is_datetime64_any_dtype(series))
                if is_text and series.notna().any():
                    unregistered.add(f"{sheet}.{col}")
                continue
            known = mapping["categories"].get(category, {})
            for value in df[col].dropna().unique():
                text = str(value).strip()
                if (text and text not in known
                        and text.upper() not in PUBLIC_TERMS
                        and not ARTEFACTS.match(text)):
                    unmapped.setdefault(category, set()).add(text)

    print(f"Scanned {path.name}: {len(frames)} sheet(s)\n")
    if unmapped:
        print("NEW VALUES needing a surrogate:")
        for category, values in sorted(unmapped.items()):
            preview = sorted(values)[:6]
            more = f" (+{len(values) - len(preview)} more)" if len(values) > 6 else ""
            print(f"  {category:<16} {len(values):>4} new   e.g. {preview}{more}")
    else:
        print("No new values - the mapping already covers this file.")

    if metadata_seen:
        print(f"\nMETADATA SHEETS needing manual review: {', '.join(metadata_seen)}")
        print("  These describe the data but still carry names - group and team")
        print("  names, worked examples, venue lists. Check them by hand.")

    if unregistered:
        print("\nUNREGISTERED TEXT COLUMNS - not in COLUMN_CATEGORIES, so they")
        print("would pass through UNCHANGED. Classify each before applying:")
        for name in sorted(unregistered):
            print(f"  - {name}")
        return 1
    return 0


def cmd_apply(path: Path, out: Path) -> int:
    mapping = _load()
    frames = _read(path)
    changed: dict[str, int] = {}

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet, df in frames.items():
            result = df.copy()
            if sheet in METADATA_SHEETS:
                # These sheets are key-value or list shaped, so column rules do
                # not apply - but they carry the enum domains and worked
                # examples, which is where anonymisation most often leaks. Sweep
                # every cell and replace any value already in the mapping.
                swept = _sweep_known(result, mapping)
                changed["(metadata) " + sheet] = swept
                result.to_excel(writer, sheet_name=sheet[:31], index=False)
                continue
            for col in result.columns:
                category = _category_for(col)
                if not category:
                    continue
                result[col] = result[col].map(
                    lambda v: _map_value(v, category, mapping) if pd.notna(v) else v)
                changed[col] = changed.get(col, 0) + int(result[col].notna().sum())
            result.to_excel(writer, sheet_name=sheet[:31], index=False)

    _save(mapping)
    print(f"Anonymised -> {out}")
    print(f"Mapping updated -> {MAPPING_FILE} (git-ignored)\n")
    for col, n in sorted(changed.items()):
        print(f"  {col:<32} {n:>5} values replaced")
    total = sum(len(v) for v in mapping["categories"].values())
    print(f"\nGolden source now holds {total} mappings across "
          f"{len(mapping['categories'])} categories.")
    return 0


def cmd_report(category: str | None) -> int:
    mapping = _load()
    cats = mapping["categories"]
    if not cats:
        print("Mapping is empty. Run `apply` on a real extract first.")
        return 0
    for name, table in sorted(cats.items()):
        if category and name != category:
            continue
        print(f"\n=== {name} ({len(table)} mappings) ===")
        for real, fake in sorted(table.items()):
            print(f"  {real:<44} -> {fake}")
    return 0


def cmd_audit(path: Path) -> int:
    """Look for values in an anonymised file that are still in the mapping keys,
    i.e. real values that escaped replacement."""
    mapping = _load()
    # A value deliberately mapped to itself (because every part of it is generic
    # vocabulary) is not a leak - exclude those or the audit cries wolf.
    reals = {r.casefold()
             for table in mapping["categories"].values()
             for r, fake in table.items() if r != fake}
    if not reals:
        print("Mapping is empty - nothing to audit against.")
        return 0

    leaks: list[str] = []
    for sheet, df in _read(path).items():
        for col in df.columns:
            for value in df[col].dropna().unique():
                if str(value).strip().casefold() in reals:
                    leaks.append(f"{sheet}.{col}: {value}")

    if leaks:
        print(f"LEAKED - {len(leaks)} real value(s) still present:")
        for leak in leaks[:40]:
            print(f"  {leak}")
        return 1
    print(f"Clean: no mapped real values found in {path.name}.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("scan", help="report values not yet mapped")
    p.add_argument("file", type=Path)

    p = sub.add_parser("apply", help="anonymise a file")
    p.add_argument("file", type=Path)
    p.add_argument("--out", type=Path, required=True)

    p = sub.add_parser("report", help="show the mapping")
    p.add_argument("--category")

    p = sub.add_parser("audit", help="check a file for unreplaced real values")
    p.add_argument("file", type=Path)

    args = parser.parse_args()
    if args.command == "scan":
        return cmd_scan(args.file)
    if args.command == "apply":
        return cmd_apply(args.file, args.out)
    if args.command == "report":
        return cmd_report(args.category)
    if args.command == "audit":
        return cmd_audit(args.file)
    return 1


if __name__ == "__main__":
    sys.exit(main())
